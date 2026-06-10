import json
import pathlib
import subprocess
import sys

import pytest
from openpyxl import load_workbook

import build_evidence_appendix as bea

ROOT = pathlib.Path(__file__).resolve().parent.parent
SCRIPT = ROOT / "skills" / "scope-calculator" / "scripts" / "build_evidence_appendix.py"

DATA = {
    "customer": "Acme Corp", "date": "2026-06-07",
    "rollup": {"url_total": 410_000, "path_floor": 91_000, "spiral_adjusted_anchor": 95_000,
               "low": 88_000, "high": 105_000, "confidence": "MEDIUM",
               "census_ids": [812], "crawl_status": "done"},
    "per_domain": [
        {"hostname": "acme.com", "raw_urls": 90_000, "paths": 86_000, "spiral_flag": False,
         "spiral_ratio": 1.05, "defensible_pages": 90_000, "discounted": 0, "why": "",
         "url_samples": ["https://acme.com/", "https://acme.com/about"]},
        {"hostname": "shop.acme.com", "raw_urls": 320_000, "paths": 5_000, "spiral_flag": True,
         "spiral_ratio": 64.0, "defensible_pages": 5_000, "discounted": 315_000,
         "why": "64x query-param spiral", "url_samples": ["https://shop.acme.com/p/1"]}],
    "usage": {"consent_states": {"count": 3, "names": ["Default", "Opt-Out", "GPC"]},
              "pages_per_sweep": 285_000, "annual_scans": 427_500,
              "recommended_price": 54_000, "recommended_scans": 430_167,
              "cadence_layers": [
                  {"name": "Full privacy sweep", "runs_per_year": 1, "pct": 1.0, "pages": 285_000, "runs": 285_000},
                  {"name": "Priority pages", "runs_per_year": 4, "pct": 0.05, "pages": 14_250, "runs": 57_000},
                  {"name": "Consent-critical pages", "runs_per_year": 12, "pct": 0.025, "pages": 7_125, "runs": 85_500}]},
}


def _alltext(wb):
    out = []
    for ws in wb.worksheets:
        out.append(ws.title)
        for row in ws.iter_rows(values_only=True):
            out += [str(v) for v in row if v is not None]
    return "\n".join(out)


def test_invariant_raises_on_mismatch():
    bad = json.loads(json.dumps(DATA))
    bad["rollup"]["spiral_adjusted_anchor"] = 9_999
    with pytest.raises(ValueError, match=r"sum \d+ != rollup anchor"):
        bea.build_workbook(bad)


def test_sheets_present_with_usage():
    wb = bea.build_workbook(DATA)
    assert wb.sheetnames == ["Scope Summary", "Pages by Domain", "Sample Pages",
                             "Annual Usage Breakdown", "Methodology"]


def test_usage_sheet_omitted_when_no_usage():
    d = json.loads(json.dumps(DATA))
    d.pop("usage")
    wb = bea.build_workbook(d)
    assert "Annual Usage Breakdown" not in wb.sheetnames
    assert len(wb.sheetnames) == 4


def test_pages_by_domain_headers_and_fillable():
    ws = bea.build_workbook(DATA)["Pages by Domain"]
    assert [c.value for c in ws[3]] == ["Property (domain)", "Real pages", "% of total",
                                        "Spiral?"] + bea.FILL_COLS
    assert ws.cell(4, 1).value == "acme.com"      # sorted by pages desc
    assert ws.cell(4, 2).value == 90_000
    # customer-fillable columns present and empty
    assert ws.cell(4, 5).value is None and ws.cell(4, 6).value is None and ws.cell(4, 7).value is None


def test_sample_pages_present():
    t = _alltext(bea.build_workbook(DATA))
    assert "https://acme.com/about" in t
    assert "https://shop.acme.com/p/1" in t


def test_usage_breakdown_content():
    t = _alltext(bea.build_workbook(DATA))
    assert "Default" in t and "Opt-Out" in t and "GPC" in t      # consent states
    assert "TOTAL ANNUAL PAGE SCANS" in t
    assert "427500" in t                                          # annual scans total (raw value)
    assert "Annually" in t and "Quarterly" in t and "Monthly" in t  # cadence frequencies
    assert "5.0%" in t and "2.5%" in t                            # % of pages
    assert "430,167 page scans" in t and "$54,000 / year" in t    # reconciling contract sentence


def test_methodology_shows_reduction():
    t = _alltext(bea.build_workbook(DATA))
    assert "64x query-param spiral" in t
    assert "320000" in t                                          # raw URLs for the spiral domain


def test_cli_writes_file(tmp_path):
    f = tmp_path / "in.json"; f.write_text(json.dumps(DATA))
    out = tmp_path / "evidence.xlsx"
    res = subprocess.run([sys.executable, str(SCRIPT), str(f), str(out)],
                         capture_output=True, text=True)
    assert res.returncode == 0, res.stderr
    assert res.stdout.strip() == str(out)
    assert "Sample Pages" in load_workbook(out).sheetnames
