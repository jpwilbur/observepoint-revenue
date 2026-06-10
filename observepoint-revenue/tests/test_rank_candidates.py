import json
import pathlib
import subprocess
import sys

import rank_candidates as rc
import score_account as sa

ROOT = pathlib.Path(__file__).resolve().parent.parent
SCRIPT = ROOT / "skills" / "find-accounts" / "scripts" / "rank_candidates.py"
CONFIG_PATH = ROOT / "skills" / "research-account" / "references" / "scoring-config.json"
CONFIG = json.loads(CONFIG_PATH.read_text())

AS_OF = "2026-06-09"


def _data(candidates, date=AS_OF):
    return {"territory": {"region": "US West", "verticals": ["healthcare"]},
            "prepared_by": "Test", "date": date, "requested": 5, "candidates": candidates}


def _cand(name, key="pixelWiretapSuit", date="2026-06-01", url="https://example.org/x", **kw):
    c = {"name": name, "vertical": "healthcare", "reason": f"{name} reason",
         "triggerKey": key, "triggerDate": date, "sourceUrl": url}
    c.update(kw)
    return c


def test_recency_decay_parity_with_score_account():
    now_ms = sa._parse_ms(AS_OF)
    for d in ["2026-05-20", "2026-03-09", "2025-06-09", "2025-03-16", "2024-01-15",
              "2023-12-01", None, "2026", "2025-08"]:
        assert rc.recency_factor(d, CONFIG["recency"], now_ms) == \
            sa.recency_factor(d, CONFIG["recency"], now_ms), f"decay mismatch for {d!r}"


def test_round_half_up_matches_js_not_bankers():
    assert rc._round_half_up(2.5) == 3        # Python round(2.5) == 2 — we need JS Math.round
    assert rc._round_half_up(12.5) == 13
    assert rc._round_half_up(12.4) == 12


def test_half_up_rounding_applied_to_points():
    # enforcementAction = 25 pts; 2025-03-16 is exactly 450 days = 15.0 months before 2026-06-09
    # → factor 1 - (15-6)/18 = 0.5 → 12.5 → half-up 13 (banker's would give 12).
    ranked, dropped, new = rc.rank(_data([_cand("A", key="enforcementAction",
                                                date="2025-03-16")]), CONFIG)
    assert ranked[0]["effectivePoints"] == 13


def test_ranks_by_effective_points_decay_can_flip_order():
    # settlement 15 pts recent (→15) beats pixelWiretapSuit 30 pts from 2023 (>24mo → ×0.1 → 3).
    ranked, dropped, new = rc.rank(_data([
        _cand("OldSuit Co", key="pixelWiretapSuit", date="2023-01-01"),
        _cand("FreshSettle Co", key="settlement", date="2026-06-01"),
    ]), CONFIG)
    assert [e["name"] for e in ranked] == ["FreshSettle Co", "OldSuit Co"]
    assert ranked[0]["effectivePoints"] == 15 and ranked[1]["effectivePoints"] == 3


def test_tiebreak_newer_first_undated_last():
    # All pixelWiretapSuit (30): A 2026-06-01 (30), B 2026-05-01 (30); C 2025-05-09 is 13.2mo
    # → ×0.6 → 18, D undated → ×0.6 → 18. Expect A, B (newer dated first), then C before D.
    ranked, dropped, new = rc.rank(_data([
        _cand("D", date=None), _cand("C", date="2025-05-09"),
        _cand("B", date="2026-05-01"), _cand("A", date="2026-06-01"),
    ]), CONFIG)
    assert [e["name"] for e in ranked] == ["A", "B", "C", "D"]


def test_unknown_trigger_key_is_hard_error():
    import pytest
    with pytest.raises(ValueError, match="bipaOnly"):
        rc.rank(_data([_cand("X", key="bipaOnly")]), CONFIG)


def test_missing_source_url_is_hard_error():
    import pytest
    with pytest.raises(ValueError, match="sourceUrl"):
        rc.rank(_data([_cand("X", url="  ")]), CONFIG)


def test_trigger_label_attached_from_config():
    ranked, _, _ = rc.rank(_data([_cand("A")]), CONFIG)
    assert ranked[0]["triggerLabel"] == CONFIG["whyNow"]["pixelWiretapSuit"]["label"]


# ── Task 2: Seen-log ──────────────────────────────────────────────────────────

SEEN = {"candidates": [{"name": "The Example Health-System, Inc.", "firstSeen": "2026-05-01",
                        "triggerKey": "pixelWiretapSuit", "sourceUrl": "https://example.org/old"}]}


def test_seen_dedup_normalized_name_match():
    ranked, dropped, new = rc.rank(_data([_cand("Example Health System Inc"),
                                          _cand("Brand New Co")]), CONFIG, seen=SEEN)
    assert [e["name"] for e in ranked] == ["Brand New Co"]
    assert dropped == 1
    assert [n["name"] for n in new] == ["Brand New Co"]   # only NEW names get log entries


def test_include_seen_keeps_and_annotates_without_relogging():
    ranked, dropped, new = rc.rank(_data([_cand("Example Health System Inc")]), CONFIG,
                                   seen=SEEN, include_seen=True)
    assert dropped == 0
    assert ranked[0]["firstSeen"] == "2026-05-01"
    assert new == []                                       # seen entry NOT re-appended


def test_new_entry_shape_uses_data_date():
    _, _, new = rc.rank(_data([_cand("Brand New Co")]), CONFIG, seen=SEEN)
    assert new == [{"name": "Brand New Co", "firstSeen": AS_OF,
                    "triggerKey": "pixelWiretapSuit", "sourceUrl": "https://example.org/x"}]


def test_load_seen_missing_and_corrupt_treated_empty(tmp_path):
    assert rc.load_seen(tmp_path / "nope.json") == {"candidates": []}
    bad = tmp_path / "bad.json"; bad.write_text("{not json")
    assert rc.load_seen(bad) == {"candidates": []}
    wrong = tmp_path / "wrong.json"; wrong.write_text(json.dumps({"candidates": "oops"}))
    assert rc.load_seen(wrong) == {"candidates": []}
    assert rc.load_seen(None) == {"candidates": []}


# ── Task 3: Chat output + CLI ─────────────────────────────────────────────────

def _run_cli(tmp_path, data, *extra):
    f = tmp_path / "cands.json"; f.write_text(json.dumps(data))
    return subprocess.run([sys.executable, str(SCRIPT), str(f), str(CONFIG_PATH), *extra],
                          capture_output=True, text=True)


def test_cli_prints_ranked_order_and_sources(tmp_path):
    res = _run_cli(tmp_path, _data([_cand("Beta Co", key="settlement"),
                                    _cand("Alpha Co", key="pixelWiretapSuit")]))
    assert res.returncode == 0, res.stderr
    out = res.stdout
    assert out.index("1. Alpha Co") < out.index("2. Beta Co")   # 30 pts before 15 pts
    assert "https://example.org/x" in out
    assert CONFIG["whyNow"]["pixelWiretapSuit"]["label"] in out


def test_cli_seen_log_appends_and_excludes(tmp_path):
    seen = tmp_path / "seen.json"
    res1 = _run_cli(tmp_path, _data([_cand("Alpha Co")]), "--seen", str(seen))
    assert res1.returncode == 0, res1.stderr
    logged = json.loads(seen.read_text())["candidates"]
    assert [(c["name"], c["firstSeen"]) for c in logged] == [("Alpha Co", AS_OF)]
    # second run: Alpha excluded with a note; Beta appended.
    res2 = _run_cli(tmp_path, _data([_cand("Alpha Co"), _cand("Beta Co")]), "--seen", str(seen))
    assert "Beta Co" in res2.stdout and "1 previously-seen" in res2.stdout
    assert "1. Alpha Co" not in res2.stdout
    names = [c["name"] for c in json.loads(seen.read_text())["candidates"]]
    assert names == ["Alpha Co", "Beta Co"]


def test_cli_include_seen_shows_annotation(tmp_path):
    seen = tmp_path / "seen.json"
    _run_cli(tmp_path, _data([_cand("Alpha Co")]), "--seen", str(seen))
    res = _run_cli(tmp_path, _data([_cand("Alpha Co")]), "--seen", str(seen), "--include-seen")
    assert "Alpha Co" in res.stdout and f"[seen {AS_OF}]" in res.stdout


def test_cli_validation_error_exits_nonzero(tmp_path):
    res = _run_cli(tmp_path, _data([_cand("X", key="notAKey")]))
    assert res.returncode != 0
    assert "notAKey" in res.stderr


def test_cli_empty_candidates_is_valid(tmp_path):
    res = _run_cli(tmp_path, _data([]))
    assert res.returncode == 0, res.stderr
    assert "Ranked candidates (0)" in res.stdout


def test_cli_writes_no_files_without_xlsx_flag(tmp_path):
    _run_cli(tmp_path, _data([_cand("Alpha Co")]))
    assert list(tmp_path.glob("*.xlsx")) == [] and list(tmp_path.glob("*.txt")) == []


# ── Task 4: --xlsx discovery radar ───────────────────────────────────────────

def test_xlsx_radar_columns_hyperlink_fillable(tmp_path):
    from openpyxl import load_workbook
    out = tmp_path / "radar.xlsx"
    res = _run_cli(tmp_path, _data([_cand("Alpha Co")]), "--xlsx", str(out))
    assert res.returncode == 0, res.stderr
    assert out.exists() and str(out) in res.stdout
    wb = load_workbook(out)
    ws = wb["Discovery radar"]
    hdr = [c.value for c in ws[1]]
    assert hdr == ["Rank", "Company", "Vertical", "Trigger", "Trigger date", "Why now",
                   "Source", "Pursue?", "Notes"]
    assert ws.cell(row=2, column=1).value == 1
    assert ws.cell(row=2, column=2).value == "Alpha Co"
    src = ws.cell(row=2, column=7)
    assert src.hyperlink is not None and src.hyperlink.target == "https://example.org/x"
    assert ws.cell(row=2, column=8).value in (None, "")    # Pursue? fillable
    assert ws.cell(row=2, column=9).value in (None, "")    # Notes fillable


def test_xlsx_rows_follow_rank_order(tmp_path):
    from openpyxl import load_workbook
    out = tmp_path / "radar.xlsx"
    _run_cli(tmp_path, _data([_cand("Beta Co", key="settlement"),
                              _cand("Alpha Co", key="pixelWiretapSuit")]), "--xlsx", str(out))
    ws = load_workbook(out)["Discovery radar"]
    assert [(r[0], r[1]) for r in ws.iter_rows(min_row=2, values_only=True)] == \
        [(1, "Alpha Co"), (2, "Beta Co")]


def test_normalize_name_word_level_article_strip():
    # leading standalone 'The' is dropped...
    assert rc.normalize_name("The Example Health-System, Inc.") == \
        rc.normalize_name("Example Health System Inc")
    # ...but names that merely START with 'the' as one word are untouched.
    assert rc.normalize_name("Theranos") == "theranos"
    assert rc.normalize_name("Theranos") != rc.normalize_name("Ranos")
