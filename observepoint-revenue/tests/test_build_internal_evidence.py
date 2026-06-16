# tests/test_build_internal_evidence.py
import json
import pathlib
import subprocess
import sys

import pytest
from openpyxl import load_workbook

import build_internal_evidence as bie

ROOT = pathlib.Path(__file__).resolve().parent.parent
SCRIPT = ROOT / "skills" / "scope-calculator" / "scripts" / "build_internal_evidence.py"

DATA = {
    "customer": "Acme Corp", "date": "2026-06-15",
    "rollup": {"spiral_adjusted_anchor": 95_000, "low": 88_000, "high": 105_000,
               "confidence": "MEDIUM", "census_ids": [812], "crawl_status": "done"},
    "per_domain": [
        {"hostname": "acme.com", "raw_urls": 90_000, "defensible_pages": 90_000, "discounted": 0, "why": ""},
        {"hostname": "shop.acme.com", "raw_urls": 320_000, "defensible_pages": 5_000, "discounted": 315_000,
         "spiral_flag": True, "why": "64x query-param spiral"}],
    "pricing": {"price_by_band": [{"band_limit": 1000, "rate": 0.0, "pages": 1000, "cost": 0.0},
                                  {"band_limit": 50000, "rate": 0.17, "pages": 50000, "cost": 8500.0}],
                "modeled_scans": 430_744, "modeled_price": 54_069.28,
                "recommended_scans": 430_167, "recommended_price": 54_000,
                "pricing_source": "live @ https://app.observepoint.com/www-pricing/main.js"},
    "internal": {"assumptions": ["Geographies defaulted to 1 — confirm regions.",
                                 "Consent states assumed CCPA (3) — confirm regulations."],
                 "implied_frequency": 1.5},
}


def _alltext(wb):
    out = []
    for ws in wb.worksheets:
        out.append(ws.title)
        for row in ws.iter_rows(values_only=True):
            out += [str(v) for v in row if v is not None]
    return "\n".join(out)


def test_invariant_raises_on_mismatch():
    bad = json.loads(json.dumps(DATA))
    bad["rollup"]["spiral_adjusted_anchor"] = 9_999
    with pytest.raises(ValueError, match=r"sum \d+ != rollup anchor"):
        bie.build_workbook(bad)


def test_internal_content_present():
    t = _alltext(bie.build_workbook(DATA))
    assert "812" in t                              # census id (internal)
    assert "410,000" in t or "410000" in t         # raw URL total (internal)
    assert "64x query-param spiral" in t           # per-domain derivation note
    assert "confirm regulations" in t.lower()      # assumption-to-verify
    assert "430,744" in t or "430744" in t         # modeled scans


def test_dominance_flag_when_one_domain_dominates():
    # spec §9.2: flag when the single largest host exceeds ~40% of the anchor.
    d = json.loads(json.dumps(DATA))
    d["per_domain"] = [{"hostname": "trap.com", "raw_urls": 1_710_000, "defensible_pages": 90_000,
                        "discounted": 0, "why": ""},
                       {"hostname": "acme.com", "raw_urls": 5_000, "defensible_pages": 5_000,
                        "discounted": 0, "why": ""}]
    d["rollup"]["spiral_adjusted_anchor"] = 95_000   # trap.com = 94.7% of anchor
    t = _alltext(bie.build_workbook(d))
    assert "trap.com" in t and "%" in t
    assert bie.dominant_host(d) is not None and bie.dominant_host(d)["hostname"] == "trap.com"


def test_no_dominance_flag_when_balanced():
    # Use a 3-domain balanced split: 35k/30k/30k = 95k.
    # Largest share = 35k/95k = 36.8% which is < 40% threshold → no dominance flag.
    d = json.loads(json.dumps(DATA))
    d["per_domain"] = [
        {"hostname": "a.com", "raw_urls": 36_000, "defensible_pages": 35_000, "discounted": 1_000, "why": ""},
        {"hostname": "b.com", "raw_urls": 30_000, "defensible_pages": 30_000, "discounted": 0, "why": ""},
        {"hostname": "c.com", "raw_urls": 30_000, "defensible_pages": 30_000, "discounted": 0, "why": ""},
    ]
    d["rollup"]["spiral_adjusted_anchor"] = 95_000   # 35k+30k+30k = 95k; max share 36.8% < 40%
    assert bie.dominant_host(d) is None


def test_sheets_present():
    wb = bie.build_workbook(DATA)
    titles = wb.sheetnames
    assert "Derivation (INTERNAL)" in titles
    assert "Per-Domain (INTERNAL)" in titles
    assert "Assumptions (INTERNAL)" in titles
    assert "Pricing (INTERNAL)" in titles


def test_derivation_sheet_has_census_and_confidence():
    wb = bie.build_workbook(DATA)
    ws = wb["Derivation (INTERNAL)"]
    vals = [str(v) for row in ws.iter_rows(values_only=True) for v in row if v is not None]
    flat = " ".join(vals)
    assert "812" in flat
    assert "MEDIUM" in flat
    assert "done" in flat


def test_per_domain_sheet_has_spiral_note():
    wb = bie.build_workbook(DATA)
    ws = wb["Per-Domain (INTERNAL)"]
    vals = [str(v) for row in ws.iter_rows(values_only=True) for v in row if v is not None]
    flat = " ".join(vals)
    assert "64x query-param spiral" in flat
    assert "shop.acme.com" in flat


def test_pricing_sheet_has_band_detail():
    wb = bie.build_workbook(DATA)
    ws = wb["Pricing (INTERNAL)"]
    vals = [str(v) for row in ws.iter_rows(values_only=True) for v in row if v is not None]
    flat = " ".join(vals)
    assert "0.17" in flat   # band rate
    assert "live @ https://app.observepoint.com" in flat  # pricing source


def test_cli_writes_file(tmp_path):
    inp = tmp_path / "inp.json"
    out = tmp_path / "out.xlsx"
    inp.write_text(json.dumps(DATA))
    result = subprocess.run(
        [sys.executable, str(SCRIPT), str(inp), str(out)],
        capture_output=True, text=True
    )
    assert result.returncode == 0, result.stderr
    assert out.exists()
    wb = load_workbook(str(out))
    assert "Derivation (INTERNAL)" in wb.sheetnames


def test_cli_bad_json_exits_nonzero(tmp_path):
    inp = tmp_path / "bad.json"
    inp.write_text("not json")
    result = subprocess.run(
        [sys.executable, str(SCRIPT), str(inp), str(tmp_path / "out.xlsx")],
        capture_output=True, text=True
    )
    assert result.returncode != 0


def test_pricing_floats_have_decimal_format():
    """Band rate 0.17 must use a decimal number format, not the integer '#,##0'."""
    wb = bie.build_workbook(DATA)
    ws = wb["Pricing (INTERNAL)"]
    rate_cell = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == 50000:
                # The rate cell is immediately to the right (same row, next column)
                rate_cell = ws.cell(cell.row, cell.column + 1)
                break
        if rate_cell is not None:
            break
    assert rate_cell is not None, "Could not locate band-rate cell (row where first cell is 50000)"
    assert rate_cell.value == 0.17
    assert "." in rate_cell.number_format, (
        f"Expected a decimal format (containing '.') for float 0.17, got {rate_cell.number_format!r}"
    )
    assert rate_cell.number_format == "#,##0.##"
