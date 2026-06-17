# tests/test_read_scope_of_work.py
"""TDD suite for read_scope_of_work.py — written RED-first (module absent), then green.

Fixture: a realistic workbook built fresh from build_model.build_workbook with
  ~23 domains (so the aggregate row exists), multipliers {1,3,1}, 4-layer ladder,
  buffer_pct=0.15, BAKED_TIERS. All tests build from scratch into tmp_path so
  the suite is hermetic.
"""
import copy
import json
import pathlib
import subprocess
import sys

import pytest

import build_model as bm
import compute_scope as cs
import build_proposal as bp

# ---------------------------------------------------------------------------
# Shared fixture data
# ---------------------------------------------------------------------------

TIERS = cs.BAKED_TIERS

LAYERS = [
    {"name": "Baseline inventory",
     "why": "A full sweep so nothing is invisible.",
     "pct": 1.0, "runs_per_year": 1},
    {"name": "High Priority",
     "why": "Aligned to release cadence.",
     "pct": 0.015, "runs_per_year": 52},
    {"name": "Moderate Priority Pages",
     "why": "Monthly audit of the main body.",
     "pct": 0.075, "runs_per_year": 12},
    {"name": "Low Priority Pages",
     "why": "Quarterly sweep of the long tail.",
     "pct": 0.20, "runs_per_year": 4},
]

# 23 domains → 20 individual + 1 aggregate row (3 tail)
_PAGES = [9000, 8000, 7000, 6000, 5000, 4000, 3000, 2000, 1000, 900,
          800,  700,  600,  500,  400,  300,  200,  100,   90,  80,
          70,    60,   50]   # 23 entries

PER_DOMAIN = [
    {"hostname": f"d{i}.com", "defensible_pages": p,
     "url_samples": [f"https://d{i}.com/"]}
    for i, p in enumerate(_PAGES)
]

BASE = sum(_PAGES)   # total defensible pages (all in-scope at 100% sample) = Total Pages Found

MODEL_DATA = {
    "customer": "Acme Corp",
    "date": "2026-06-16",
    "page_count": {"low": BASE - 5_000, "anchor": BASE, "high": BASE + 7_000},
    "multipliers": {"geographies": 1, "scenarios": 3, "environments": 1},
    "cadence_layers": LAYERS,
    "buffer_pct": 0.15,
    "tiers": TIERS,
    "per_domain": PER_DOMAIN,
}


def _build_wb(tmp_path, data=None):
    """Build a fresh workbook, save to tmp_path, return the saved path."""
    d = copy.deepcopy(data or MODEL_DATA)
    wb = bm.build_workbook(d)
    path = tmp_path / "Acme Corp - Scope of Work.xlsx"
    wb.save(str(path))
    return path


# ---------------------------------------------------------------------------
# Test 1: roundtrip reader extracts inputs correctly
# ---------------------------------------------------------------------------

def test_roundtrip_reader_extracts_inputs(tmp_path):
    """Build → read → verify per_domain, multipliers, cadence_layers, buffer_pct, tiers
    all match the model inputs; page_count.anchor == compute_scope.total_pages_found(per_domain)."""
    import read_scope_of_work as rsow

    path = _build_wb(tmp_path)
    result = rsow.read_scope_of_work(str(path))

    # Multipliers
    assert result["multipliers"]["geographies"] == 1
    assert result["multipliers"]["scenarios"] == 3
    assert result["multipliers"]["environments"] == 1

    # Buffer
    assert abs(result["buffer_pct"] - 0.15) < 1e-6

    # Cadence layers: 4 layers
    assert len(result["cadence_layers"]) == 4
    assert result["cadence_layers"][0]["name"] == "Baseline inventory"
    assert result["cadence_layers"][0]["pct"] == pytest.approx(1.0)
    assert result["cadence_layers"][0]["runs_per_year"] == 1
    assert result["cadence_layers"][1]["runs_per_year"] == 52
    assert result["cadence_layers"][2]["runs_per_year"] == 12
    assert result["cadence_layers"][3]["runs_per_year"] == 4

    # per_domain: 21 entries (20 individual + 1 aggregate)
    assert len(result["per_domain"]) == 21

    # Total pages found matches compute_scope.total_pages_found (all in-scope at 1.0 sample = BASE)
    expected_tpf = cs.total_pages_found(result["per_domain"])
    assert result["page_count"]["anchor"] == expected_tpf
    assert result["page_count"]["low"] == expected_tpf
    assert result["page_count"]["high"] == expected_tpf

    # Tiers: 6 bands matching BAKED_TIERS
    assert len(result["tiers"]) == len(TIERS)
    assert result["tiers"][0]["pricePerPage"] == pytest.approx(TIERS[0]["pricePerPage"])
    assert result["tiers"][1]["pricePerPage"] == pytest.approx(TIERS[1]["pricePerPage"])


# ---------------------------------------------------------------------------
# Test 2: validator passes a clean workbook
# ---------------------------------------------------------------------------

def test_validator_passes_clean_workbook(tmp_path):
    """A fresh workbook built by build_model must produce zero hard_stops."""
    import read_scope_of_work as rsow
    from openpyxl import load_workbook

    path = _build_wb(tmp_path)
    wb_f = load_workbook(str(path), data_only=False)
    wb_v = load_workbook(str(path), data_only=True)

    # Discover geometry
    geom = rsow._discover_geometry(wb_f)
    hard_stops, warnings = rsow.validate_scope_of_work(wb_f, wb_v, geom=geom)
    assert hard_stops == [], f"Unexpected hard_stops: {hard_stops}"


# ---------------------------------------------------------------------------
# Test 3: payload feeds build_proposal successfully
# ---------------------------------------------------------------------------

def test_payload_feeds_build_proposal(tmp_path):
    """proposal_payload_from_scope_of_work → build_proposal.build_proposal renders without
    raising; the document text contains the graduated price formatted as $X."""
    import read_scope_of_work as rsow

    path = _build_wb(tmp_path)
    payload = rsow.proposal_payload_from_scope_of_work(
        str(path),
        prepared_by="Jarrod Wilbur",
        date="2026-06-16",
        use_case="privacy",
        regulations=["GDPR", "CCPA"],
        monitoring_summary="ObservePoint will monitor all in-scope properties.",
        properties_note=None,
        consent_names=["Default", "Opt-out", "GPC"],
    )

    # Must have the required keys
    assert "combined_pages" in payload["usage"]
    assert "predicted_scans" in payload["usage"]
    assert "predicted_price" in payload["pricing"]

    # build_proposal must not raise
    doc = bp.build_proposal(payload)
    assert doc is not None

    # Document text should contain the graduated price
    full_text = "\n".join(
        p.text for p in doc.paragraphs
    )
    # Also check tables
    for table in doc.tables:
        for row in table.rows:
            for cell in row.cells:
                full_text += "\n" + cell.text

    predicted_scans = payload["usage"]["predicted_scans"]
    expected_price = cs.graduated_price(predicted_scans, cs.BAKED_TIERS)["total"]
    expected_str = bp._usd(expected_price)
    assert expected_str in full_text, (
        f"Expected '{expected_str}' in doc text. predicted_scans={predicted_scans}"
    )


# ---------------------------------------------------------------------------
# Test 4: tampered formula (literal overwrite) → hard stop
# ---------------------------------------------------------------------------

def test_tampered_formula_hard_stops(tmp_path):
    """Overwrite Scope of Work!B6 with a literal integer → IntegrityError naming the cell."""
    import read_scope_of_work as rsow
    from openpyxl import load_workbook
    from openpyxl.worksheet.formula import ArrayFormula

    path = _build_wb(tmp_path)

    # Overwrite B6 with a literal
    wb = load_workbook(str(path))
    ws = wb["Scope of Work"]
    ws["B6"] = 99999   # literal, not a formula
    tampered = tmp_path / "tampered.xlsx"
    wb.save(str(tampered))

    with pytest.raises(rsow.IntegrityError) as exc_info:
        rsow.read_scope_of_work(str(tampered))

    msg = str(exc_info.value)
    assert "B6" in msg or "Total Pages Found" in msg, f"Expected B6 in error: {msg}"
    assert "scope-calculator:" in msg


def test_tampered_formula_b10_hard_stops(tmp_path):
    """Overwrite Scope of Work!B10 with a literal → IntegrityError."""
    import read_scope_of_work as rsow
    from openpyxl import load_workbook

    path = _build_wb(tmp_path)
    wb = load_workbook(str(path))
    ws = wb["Scope of Work"]
    ws["B10"] = 123456   # literal
    tampered = tmp_path / "tampered_b10.xlsx"
    wb.save(str(tampered))

    with pytest.raises(rsow.IntegrityError) as exc_info:
        rsow.read_scope_of_work(str(tampered))

    msg = str(exc_info.value)
    assert "B10" in msg or "Combined" in msg, f"Expected B10 in error: {msg}"


# ---------------------------------------------------------------------------
# Test 5: deleted structural row → hard stop
# ---------------------------------------------------------------------------

def test_deleted_structural_row_hard_stops(tmp_path):
    """Remove 'Buffer %' row by building a workbook without it, or tamper the label → IntegrityError."""
    import read_scope_of_work as rsow
    from openpyxl import load_workbook

    path = _build_wb(tmp_path)

    # Overwrite the Buffer % label so the row can't be found
    wb = load_workbook(str(path))
    ws = wb["Scope of Work"]
    # Find the buffer row by scanning column A from row 14
    for r in range(14, 40):
        if ws.cell(r, 1).value == "Buffer %":
            ws.cell(r, 1).value = "DELETED"   # corrupt the label
            break
    else:
        pytest.skip("Could not locate Buffer % row to tamper")
    tampered = tmp_path / "tampered_buf.xlsx"
    wb.save(str(tampered))

    with pytest.raises(rsow.IntegrityError) as exc_info:
        rsow.read_scope_of_work(str(tampered))

    msg = str(exc_info.value)
    assert "Buffer" in msg or "buffer" in msg or "scope-calculator:" in msg


# ---------------------------------------------------------------------------
# Test 6: invalid input → hard stop
# ---------------------------------------------------------------------------

def test_invalid_input_hard_stops(tmp_path):
    """Sample Size > 100 (e.g. 150) → IntegrityError naming the row."""
    import read_scope_of_work as rsow
    from openpyxl import load_workbook

    path = _build_wb(tmp_path)
    wb = load_workbook(str(path))
    ws = wb["Scope Detail"]
    ws["E4"] = 150   # 150 > 100 → hard-stop
    tampered = tmp_path / "tampered_sample.xlsx"
    wb.save(str(tampered))

    with pytest.raises(rsow.IntegrityError) as exc_info:
        rsow.read_scope_of_work(str(tampered))

    msg = str(exc_info.value)
    assert "Sample Size" in msg or "sample" in msg.lower(), f"Expected sample mention: {msg}"


def test_invalid_pages_hard_stops(tmp_path):
    """Negative pages value → IntegrityError naming the row."""
    import read_scope_of_work as rsow
    from openpyxl import load_workbook

    path = _build_wb(tmp_path)
    wb = load_workbook(str(path))
    ws = wb["Scope Detail"]
    ws["B4"] = -5   # negative pages
    tampered = tmp_path / "tampered_pages.xlsx"
    wb.save(str(tampered))

    with pytest.raises(rsow.IntegrityError) as exc_info:
        rsow.read_scope_of_work(str(tampered))

    msg = str(exc_info.value)
    assert "Pages" in msg or "pages" in msg.lower() or "scope-calculator:" in msg


def test_invalid_multiplier_hard_stops(tmp_path):
    """Multiplier set to 0 → IntegrityError."""
    import read_scope_of_work as rsow
    from openpyxl import load_workbook

    path = _build_wb(tmp_path)
    wb = load_workbook(str(path))
    ws = wb["Scope of Work"]
    ws["B8"] = 0   # scenarios = 0, must be ≥ 1
    tampered = tmp_path / "tampered_mult.xlsx"
    wb.save(str(tampered))

    with pytest.raises(rsow.IntegrityError) as exc_info:
        rsow.read_scope_of_work(str(tampered))

    msg = str(exc_info.value)
    assert "scenario" in msg.lower() or "B8" in msg or "multiplier" in msg.lower()


# ---------------------------------------------------------------------------
# Test 7: no in-scope pages → hard stop
# ---------------------------------------------------------------------------

def test_no_in_scope_pages_hard_stops(tmp_path):
    """All Include? = FALSE → IntegrityError (no in-scope pages)."""
    import read_scope_of_work as rsow
    from openpyxl import load_workbook

    path = _build_wb(tmp_path)
    wb = load_workbook(str(path))
    ws = wb["Scope Detail"]
    # Set all Include? cells (column D) to False
    for row in ws.iter_rows(min_row=4, max_col=4, min_col=4):
        for cell in row:
            if cell.value is not None:
                cell.value = False
    tampered = tmp_path / "tampered_no_include.xlsx"
    wb.save(str(tampered))

    with pytest.raises(rsow.IntegrityError) as exc_info:
        rsow.read_scope_of_work(str(tampered))

    msg = str(exc_info.value)
    assert "in-scope" in msg.lower() or "excluded" in msg.lower() or "0" in msg


# ---------------------------------------------------------------------------
# Test 8: sample size 50 normalised to 0.5 with warning
# ---------------------------------------------------------------------------

def test_sample_size_50_normalized_to_half(tmp_path):
    """Sample Size cell of 50 (bare number) → read as 0.5 + WARNING."""
    import read_scope_of_work as rsow
    from openpyxl import load_workbook

    path = _build_wb(tmp_path)
    wb = load_workbook(str(path))
    ws = wb["Scope Detail"]
    ws["E4"] = 50   # 50 → interpreted as 50% → normalize to 0.5
    normalized = tmp_path / "normalized_sample.xlsx"
    wb.save(str(normalized))

    result = rsow.read_scope_of_work(str(normalized))
    # The first domain's sample_size must be normalized to 0.5
    assert result["per_domain"][0]["sample_size"] == pytest.approx(0.5), (
        f"Expected 0.5 after normalization, got {result['per_domain'][0]['sample_size']}"
    )
    # There must be a warning mentioning the normalization
    warnings = result.get("_warnings", [])
    assert any("50" in w or "sample" in w.lower() or "percent" in w.lower() or "normaliz" in w.lower()
               for w in warnings), f"Expected normalization warning, got: {warnings}"


# ---------------------------------------------------------------------------
# Test 9: cache absent is a warning, not a stop
# ---------------------------------------------------------------------------

def test_cache_absent_is_warning_not_stop(tmp_path):
    """An openpyxl-built workbook (no Excel recalc → cached values are None for formula cells)
    produces reconciliation WARNINGs, not a hard stop. The payload is still returned."""
    import read_scope_of_work as rsow
    from openpyxl import load_workbook

    path = _build_wb(tmp_path)
    # Load with data_only=True on an openpyxl-only file → formulas return None
    wb_v = load_workbook(str(path), data_only=True)
    # Verify that the cached value for B6 is actually None (no real Excel recalc)
    cached_b6 = wb_v["Scope of Work"]["B6"].value
    if cached_b6 is not None:
        pytest.skip("File was recalculated by Excel — cannot test cache-absent path")

    # read_scope_of_work must succeed (not raise IntegrityError)
    result = rsow.read_scope_of_work(str(path))
    assert "page_count" in result
    assert result["page_count"]["anchor"] > 0

    # Warnings must mention the missing cache / reconciliation skip
    warnings = result.get("_warnings", [])
    assert any("recalcul" in w.lower() or "cache" in w.lower() or "reconcil" in w.lower()
               for w in warnings), f"Expected cache/reconcil warning, got: {warnings}"


# ---------------------------------------------------------------------------
# Test 10: CLI prints integrity report to stderr and exits non-zero
# ---------------------------------------------------------------------------

def test_cli_integrity_report_nonzero(tmp_path):
    """CLI on a tampered file exits non-zero, prints 'scope-calculator:' itemized report,
    no Python traceback visible."""
    import read_scope_of_work as rsow   # noqa: ensure import path works
    from openpyxl import load_workbook

    path = _build_wb(tmp_path)
    # Tamper B6 with a literal
    wb = load_workbook(str(path))
    ws = wb["Scope of Work"]
    ws["B6"] = 42000
    tampered = tmp_path / "cli_tampered.xlsx"
    wb.save(str(tampered))

    script = (pathlib.Path(__file__).resolve().parent.parent
              / "skills/scope-calculator/scripts/read_scope_of_work.py")

    result = subprocess.run(
        [sys.executable, str(script), str(tampered)],
        capture_output=True, text=True
    )
    assert result.returncode != 0, "Expected non-zero exit for tampered file"
    output = result.stdout + result.stderr
    assert "scope-calculator:" in output, f"Expected 'scope-calculator:' in output:\n{output}"
    # No raw Python traceback
    assert "Traceback" not in output, f"Unexpected traceback in output:\n{output}"
    # Itemized bullet point
    assert "•" in output or "failed integrity" in output.lower() or "HARD-STOP" in output.upper() or "B6" in output


# ---------------------------------------------------------------------------
# Fix 2: Regression test — _norm tolerates spreadsheet re-save variants
# ---------------------------------------------------------------------------

def test_norm_tolerates_spreadsheet_resave_variants():
    """_norm must treat locale ';' separators, '@' implicit-intersection, '_xlfn.' prefixes,
    '$' references, case differences, and sheet-quote style as equivalent."""
    import sys
    import pathlib
    sys.path.insert(0, str(
        pathlib.Path(__file__).resolve().parent.parent
        / "skills/scope-calculator/scripts"
    ))
    from read_scope_of_work import _norm

    # locale ';' separator  ↔  standard ',' separator
    assert _norm("=ROUND(F14*E14;2)") == _norm("=ROUND(F14*E14,2)")

    # '@' implicit-intersection prefix stripped
    assert _norm("=@F14") == _norm("=F14")

    # '_xlfn.' newer-function prefix stripped (any case on input)
    assert _norm("=_xlfn.SUM(E5:E9)") == _norm("=SUM(E5:E9)")

    # '$' stripped from cell references
    assert _norm("=$B$10*D14") == _norm("=B10*D14")

    # case normalisation
    assert _norm("=round(f14*e14,2)") == _norm("=ROUND(F14*E14,2)")

    # sheet-quote normalisation: 'Scope Detail'! ≡ Scope Detail!
    assert _norm("='Scope Detail'!B6") == _norm("=Scope Detail!B6")


# ---------------------------------------------------------------------------
# Fix 3: 5-layer-ladder geometry test
# ---------------------------------------------------------------------------

LAYERS_5 = [
    {"name": "Baseline inventory",
     "why": "A full sweep so nothing is invisible.",
     "pct": 1.0, "runs_per_year": 1},
    {"name": "High Priority",
     "why": "Aligned to release cadence.",
     "pct": 0.015, "runs_per_year": 52},
    {"name": "Moderate Priority Pages",
     "why": "Monthly audit of the main body.",
     "pct": 0.075, "runs_per_year": 12},
    {"name": "Low Priority Pages",
     "why": "Quarterly sweep of the long tail.",
     "pct": 0.20, "runs_per_year": 4},
    {"name": "Daily watch",
     "why": "Near-real-time alerting on critical paths.",
     "pct": 0.005, "runs_per_year": 365},
]


def test_five_layer_ladder_geometry(tmp_path):
    """Build a 5-layer workbook, read it back: no hard-stops; cadence_layers has 5 entries
    with matching names and pcts. Proves n_layers/buf discovery generalizes beyond 4 layers."""
    import read_scope_of_work as rsow

    model_5 = {
        "customer": "Five Layer Corp",
        "date": "2026-06-16",
        "page_count": {"low": BASE - 5_000, "anchor": BASE, "high": BASE + 7_000},
        "multipliers": {"geographies": 1, "scenarios": 3, "environments": 1},
        "cadence_layers": LAYERS_5,
        "buffer_pct": 0.15,
        "tiers": TIERS,
        "per_domain": PER_DOMAIN,
    }

    wb = bm.build_workbook(copy.deepcopy(model_5))
    path = tmp_path / "Five Layer Corp - Scope of Work.xlsx"
    wb.save(str(path))

    result = rsow.read_scope_of_work(str(path))

    # No hard-stops (the call would have raised IntegrityError)
    assert "cadence_layers" in result

    layers_out = result["cadence_layers"]
    assert len(layers_out) == 5, f"Expected 5 cadence layers, got {len(layers_out)}: {layers_out}"

    # Names and pcts round-trip correctly
    for i, expected in enumerate(LAYERS_5):
        assert layers_out[i]["name"] == expected["name"], (
            f"Layer {i} name mismatch: {layers_out[i]['name']!r} != {expected['name']!r}"
        )
        assert abs(layers_out[i]["pct"] - expected["pct"]) < 1e-9, (
            f"Layer {i} pct mismatch: {layers_out[i]['pct']} != {expected['pct']}"
        )


# ---------------------------------------------------------------------------
# Fix 4: Tightened clean-workbook lock test
# ---------------------------------------------------------------------------

def test_validator_passes_clean_workbook_no_unexpected_warnings(tmp_path):
    """A fresh openpyxl-built workbook must have zero hard_stops AND only the
    cache-absent/reconciliation-skip warning (no formula, structure, or input WARNINGs)."""
    import read_scope_of_work as rsow
    from openpyxl import load_workbook

    path = _build_wb(tmp_path)
    wb_f = load_workbook(str(path), data_only=False)
    wb_v = load_workbook(str(path), data_only=True)

    geom = rsow._discover_geometry(wb_f)
    hard_stops, warnings = rsow.validate_scope_of_work(wb_f, wb_v, geom=geom)

    assert hard_stops == [], f"Unexpected hard_stops: {hard_stops}"

    # Every warning must be about cache/recalc/reconciliation — not formulas or structure
    _ACCEPTABLE = ("cache", "recalc", "reconcil")
    unexpected = [
        w for w in warnings
        if not any(kw in w.lower() for kw in _ACCEPTABLE)
    ]
    assert unexpected == [], (
        f"Unexpected non-cache/recalc warnings on a clean workbook: {unexpected}"
    )
