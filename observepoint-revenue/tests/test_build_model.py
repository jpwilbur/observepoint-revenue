# tests/test_build_model.py
import json, pathlib, sys
import pytest
from openpyxl import load_workbook

import build_model as bm
import compute_scope as cs

ROOT = pathlib.Path(__file__).resolve().parent.parent
SCRIPT = ROOT / "skills" / "scope-calculator" / "scripts" / "build_model.py"

TIERS = cs.BAKED_TIERS
LAYERS = [
    {"name": "Baseline inventory", "why": "A full sweep so nothing is invisible.", "pct": 1.0, "runs_per_year": 1},
    {"name": "Inventory refresh", "why": "Quarterly keeps the picture current.", "pct": 0.50, "runs_per_year": 4},
    {"name": "Compliance audit", "why": "Monthly audit of the key body.", "pct": 0.15, "runs_per_year": 12},
    {"name": "Release catch", "why": "Weekly, aligned to releases.", "pct": 0.05, "runs_per_year": 52},
    {"name": "Critical watch", "why": "Daily on crown-jewel pages.", "pct": 0.01, "runs_per_year": 365},
]
DATA = {
    "customer": "Acme Corp", "date": "2026-06-16",
    "page_count": {"low": 88_000, "anchor": 95_000, "high": 105_000},
    "multipliers": {"geographies": 1, "scenarios": 3, "environments": 1},
    "cadence_layers": LAYERS, "buffer_pct": 0.0, "tiers": TIERS,
    "per_domain": [{"hostname": "acme.com", "defensible_pages": 95_000, "url_samples": ["https://acme.com/"]}],
    "rollup": {"spiral_adjusted_anchor": 95_000},
}


def emulate_model(base, geos, scenarios, env, layers, buffer_pct, tiers):
    sweep = base * geos * scenarios * env
    spl = [round((sweep * L["pct"]) * L["runs_per_year"], 2) for L in layers]
    predicted = round(sum(spl))
    purchased = round(predicted * (1 + buffer_pct))
    los, his, rates, lo = [], [], [], 0
    for t in tiers:
        los.append(lo); his.append(lo + t["limit"]); rates.append(t["pricePerPage"]); lo += t["limit"]
    his[-1] = 10**12
    price = round(sum(max(0, min(purchased, his[i]) - los[i]) * rates[i] for i in range(len(tiers))), 2)
    return {"sweep": sweep, "predicted": predicted, "purchased": purchased, "price": price}


def _cs_anchor(data):
    out = cs.compute({"page_count": data["page_count"], "multipliers": data["multipliers"],
                      "cadence_layers": data["cadence_layers"], "buffer_pct": data["buffer_pct"],
                      "tiers": data["tiers"]})
    return out["anchor"]


def test_input_cells_present_and_editable_values():
    ws = bm.build_workbook(DATA)["Investment Model"]
    assert ws["B6"].value == 95_000          # validated pages
    assert ws["B7"].value == 1 and ws["B8"].value == 3 and ws["B9"].value == 1
    assert ws["C14"].value == 1.0 and ws["D14"].value == 1     # first layer pct + runs/yr
    assert ws["A14"].value == "Baseline inventory" and ws["B14"].value.startswith("A full sweep")


def test_scan_formulas_are_formulas_not_values():
    ws = bm.build_workbook(DATA)["Investment Model"]
    assert ws["B10"].value == "=B6*B7*B8*B9"
    assert ws["E14"].value == "=$B$10*C14"
    assert ws["F14"].value == "=ROUND(E14*D14,2)"
    assert ws["F20"].value == "=ROUND(SUM(F14:F18),0)"
    assert ws["F21"].value == "=ROUND(F20*(1+B19),0)"


def test_emulator_matches_engine_at_anchor():
    a = _cs_anchor(DATA)
    e = emulate_model(95_000, 1, 3, 1, LAYERS, 0.0, TIERS)
    assert e["predicted"] == a["predicted_scans"]
    assert e["purchased"] == a["purchased_scans"]


def test_emulator_matches_engine_when_daily_dropped_and_quarterly_halved():
    layers = json.loads(json.dumps(LAYERS))
    layers[4]["pct"] = 0.0           # drop "Critical watch" (daily)
    layers[1]["pct"] = 0.25          # halve "Inventory refresh" (quarterly)
    d = json.loads(json.dumps(DATA)); d["cadence_layers"] = layers
    a = _cs_anchor(d)
    e = emulate_model(95_000, 1, 3, 1, layers, 0.0, TIERS)
    assert e["predicted"] == a["predicted_scans"] and e["purchased"] == a["purchased_scans"]
