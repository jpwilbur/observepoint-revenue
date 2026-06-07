"""Customer-facing, ObservePoint-themed evidence workbook (.xlsx).

The "show your work" companion to the proposal. Five sheets:
  1. Scope Summary          — headline page count, usage, recommended contract.
  2. Pages by Domain        — real pages per property (+ customer-fillable Include/Priority/Notes).
  3. Sample Pages           — actual example URLs per property ("proof it isn't junk").
  4. Annual Usage Breakdown — % of pages × cadence → page scans, summed to the proposal's total.
  5. Methodology            — why raw URL counts were reduced to real pages (defensibility).

ObservePoint theme: Montserrat, near-black #1E1E1E headers, brand yellow #F2CD14 accents,
light-gray #F2F2F2 alternating rows, embedded OP logo.

Enforces the spec §4.6 invariant: per-domain `defensible_pages` must sum to the rollup anchor.

Input: {customer, date?, rollup{...}, per_domain[{...url_samples[]...}], usage?{...}}.
`usage` (from size-and-price) is optional; when absent the Annual Usage Breakdown sheet is omitted.
"""
import json
import pathlib
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FONT = "Montserrat"
DARK, YELLOW, LIGHT, GRAY, WHITE = "1E1E1E", "F2CD14", "F2F2F2", "5C5C5C", "FFFFFF"
FILL_COLS = ["Include in scope?", "Priority", "Notes"]  # customer-fillable, left empty
LOGO = pathlib.Path(__file__).resolve().parents[2] / "scope-calculator" / "assets" / "op-logo.png"
_FREQ = {1: "Annually", 4: "Quarterly", 12: "Monthly", 26: "Bi-weekly", 52: "Weekly", 365: "Daily"}

_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_YBAR = Side(style="thick", color=YELLOW)


def _f(bold=False, color=DARK, size=10):
    return Font(name=FONT, bold=bold, color=color, size=size)


def _fill(hexc):
    return PatternFill("solid", fgColor=hexc)


def _check_invariant(data):
    total = sum(d["defensible_pages"] for d in data["per_domain"])
    anchor = data["rollup"]["spiral_adjusted_anchor"]
    if total != anchor:
        breakdown = ", ".join(f"{d['hostname']}={d['defensible_pages']}" for d in data["per_domain"])
        raise ValueError(f"per-domain defensible_pages sum {total} != rollup anchor {anchor} "
                         f"(per domain: {breakdown})")


def _title(ws, text, span, row):
    c = ws.cell(row, 1, text)
    c.font = _f(bold=True, size=15)
    for i in range(span):
        ws.cell(row, i + 1).border = Border(bottom=_YBAR)  # yellow accent rule under the title
    ws.row_dimensions[row].height = 22
    return row + 2


def _headers(ws, headers, row):
    for i, h in enumerate(headers):
        c = ws.cell(row, i + 1, h)
        c.font = _f(bold=True, color=WHITE)
        c.fill = _fill(DARK)
        c.border = _BORDER
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 26
    return row + 1


def _row(ws, row, values, *, alt=False, bold=(), fill=None):
    for i, v in enumerate(values):
        c = ws.cell(row, i + 1, v)
        c.font = _f(bold=(i in bold))
        c.border = _BORDER
        if fill:
            c.fill = _fill(fill)
        elif alt:
            c.fill = _fill(LIGHT)
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            c.number_format = "#,##0"
    return row + 1


def _widths(ws, widths):
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w


# ---------- sheets ----------
def _scope_summary(wb, data):
    ws = wb.active
    ws.title = "Scope Summary"
    _widths(ws, [40, 34, 22])
    if LOGO.exists():
        try:
            from openpyxl.drawing.image import Image as XLImage
            img = XLImage(str(LOGO))
            img.width, img.height = 200, 31
            ws.add_image(img, "A1")
        except Exception:
            pass
    ws.row_dimensions[1].height = 30
    r = _title(ws, "Page-Count Evidence", 3, 3)
    rollup = data["rollup"]
    usage = data.get("usage", {})
    raw_total = sum(d["raw_urls"] for d in data["per_domain"])
    def_total = sum(d["defensible_pages"] for d in data["per_domain"])
    pairs = [
        ("Customer", data.get("customer", "")),
        ("Date", data.get("date", "")),
        ("Census ID(s)", ", ".join(str(c) for c in rollup.get("census_ids", []))),
        ("Crawl status", rollup.get("crawl_status", "")),
        ("Confidence", rollup.get("confidence", "")),
        ("", ""),
        ("Validated pages — low", rollup.get("low", "")),
        ("Validated pages — anchor (recommended)", rollup.get("spiral_adjusted_anchor", "")),
        ("Validated pages — high", rollup.get("high", "")),
        ("", ""),
        ("Total raw URLs crawled", raw_total),
        ("Total real (defensible) pages", def_total),
        ("Reduced (query-string duplicates removed)", raw_total - def_total),
    ]
    if usage:
        pairs += [("", ""),
                  ("Recommended annual page scans", usage.get("recommended_scans", "")),
                  ("Recommended annual investment (USD)", usage.get("recommended_price", ""))]
    for label, val in pairs:
        c1 = ws.cell(r, 1, label)
        c1.font = _f(bold=True)
        c2 = ws.cell(r, 2, val)
        c2.font = _f()
        if isinstance(val, (int, float)) and not isinstance(val, bool):
            c2.number_format = "#,##0"
        if label.startswith("Recommended"):
            for col in (1, 2):
                ws.cell(r, col).fill = _fill(YELLOW)
        r += 1
    r += 1
    note = ws.cell(r, 1, "How to use this workbook: review 'Pages by Domain' and confirm the "
                         "in-scope properties (fill Include / Priority). 'Sample Pages' shows real "
                         "example pages found on each property. 'Annual Usage Breakdown' shows how "
                         "the page-scan total in the proposal is built. 'Methodology' explains why "
                         "inflated URL counts were reduced to real pages.")
    note.font = _f(color=GRAY, size=9)
    note.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 3, end_column=3)
    ws.freeze_panes = "A4"


def _pages_by_domain(wb, data):
    ws = wb.create_sheet("Pages by Domain")
    _widths(ws, [40, 16, 10, 12, 16, 12, 26])
    r = _title(ws, "Pages by Domain", 7, 1)
    anchor = data["rollup"]["spiral_adjusted_anchor"] or 1
    r = _headers(ws, ["Property (domain)", "Real pages", "% of total", "Spiral?"] + FILL_COLS, r)
    ws.freeze_panes = ws.cell(r, 1)
    for i, d in enumerate(sorted(data["per_domain"], key=lambda x: -x["defensible_pages"])):
        pct = round(100.0 * d["defensible_pages"] / anchor, 1)
        row = [d["hostname"], d["defensible_pages"], f"{pct}%",
               "Yes" if d.get("spiral_flag") else "No", None, None, None]
        nr = _row(ws, r, row, alt=(i % 2 == 1))
        r = nr


def _sample_pages(wb, data):
    ws = wb.create_sheet("Sample Pages")
    _widths(ws, [34, 70])
    r = _title(ws, "Sample Pages — real examples found on each property", 2, 1)
    r = _headers(ws, ["Property (domain)", "Example page URL"], r)
    ws.freeze_panes = ws.cell(r, 1)
    any_rows = False
    alt = False
    for d in sorted(data["per_domain"], key=lambda x: -x["defensible_pages"]):
        samples = d.get("url_samples", [])
        for s in samples:
            r = _row(ws, r, [d["hostname"], s], alt=alt)
            any_rows = True
        if samples:
            alt = not alt
    if not any_rows:
        _row(ws, r, ["(no per-URL samples captured)", ""])


def _usage_breakdown(wb, data):
    usage = data["usage"]
    cs = usage.get("consent_states", {"count": 1, "names": ["Default"]})
    ws = wb.create_sheet("Annual Usage Breakdown")
    _widths(ws, [34, 16, 12, 18, 12, 18])
    r = _title(ws, "Annual Usage Breakdown — how the page-scan total is built", 6, 1)
    # A: pages -> one full sweep
    r = _headers(ws, ["From pages to one full sweep", "", "", "", "", "Pages"], r)
    r = _row(ws, r, ["Validated real pages on your properties", "", "", "", "", usage["pages_per_sweep"] // cs["count"] if cs["count"] else usage["pages_per_sweep"]])
    r = _row(ws, r, [f"× Consent states monitored ({', '.join(cs['names'])})", "", "", "", "", f"×{cs['count']}"], alt=True)
    r = _row(ws, r, ["Pages per full sweep", "", "", "", "", usage["pages_per_sweep"]], bold=(0, 5), fill=YELLOW)
    r += 1
    # B: cadence
    r = _headers(ws, ["What's monitored", "How often", "% of pages", "Pages each run", "Runs/yr", "Page scans/yr"], r)
    for i, L in enumerate(usage.get("cadence_layers", [])):
        freq = _FREQ.get(L.get("runs_per_year"), f"{L.get('runs_per_year')}x/yr")
        pct = f"{round(L.get('pct', 0) * 100, 2)}%"
        r = _row(ws, r, [L["name"], freq, pct, round(L.get("pages", 0)),
                         L.get("runs_per_year", ""), round(L.get("runs", 0))], alt=(i % 2 == 1))
    r = _row(ws, r, ["TOTAL ANNUAL PAGE SCANS", "", "", "", "", usage["annual_scans"]], bold=(0, 5), fill=YELLOW)
    r += 1
    if usage.get("recommended_price") and usage.get("recommended_scans"):
        c = ws.cell(r, 1, f"Recommended contract: {usage['recommended_scans']:,} page scans  =  "
                          f"${usage['recommended_price']:,.0f} / year  (reconciles exactly in "
                          f"ObservePoint's published pricing calculator).")
        c.font = _f(bold=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)


def _methodology(wb, data):
    ws = wb.create_sheet("Methodology")
    _widths(ws, [34, 16, 14, 14, 36])
    r = _title(ws, "Methodology — why raw URLs were reduced to real pages", 5, 1)
    intro = ws.cell(r, 1, "A single page can appear under thousands of URLs (session IDs, tracking "
                          "parameters, form tokens). We count the distinct REAL page and remove the "
                          "duplicate query-string URLs, so the page count survives scrutiny.")
    intro.font = _f(color=GRAY, size=9)
    intro.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=5)
    r += 3
    r = _headers(ws, ["Property (domain)", "Raw URLs", "Real pages", "Reduced", "Note"], r)
    ws.freeze_panes = ws.cell(r, 1)
    for i, d in enumerate(sorted(data["per_domain"], key=lambda x: -x.get("discounted", 0))):
        note = d.get("why", "")
        if d.get("spiral_flag") and not note:
            note = "query-string duplicates removed"
        r = _row(ws, r, [d["hostname"], d["raw_urls"], d["defensible_pages"],
                         d.get("discounted", 0), note], alt=(i % 2 == 1))


def build_workbook(data):
    _check_invariant(data)
    wb = Workbook()
    _scope_summary(wb, data)
    _pages_by_domain(wb, data)
    _sample_pages(wb, data)
    if data.get("usage"):
        _usage_breakdown(wb, data)
    _methodology(wb, data)
    return wb


def main(argv):
    if len(argv) > 1:
        with open(argv[1]) as fh:
            raw = fh.read()
    else:
        raw = sys.stdin.read()
    out = argv[2] if len(argv) > 2 else "evidence-appendix.xlsx"
    build_workbook(json.loads(raw)).save(out)
    print(out)


if __name__ == "__main__":
    main(sys.argv)
