"""Render an owned-properties inventory workbook (.xlsx) + a confirmed-domains feed (deterministic).

Input: a candidates JSON (org + properties[] each with registrable/type/confidence/evidence/source/
host_count/sample_hosts/[all_hosts_file] + an excluded[] list). Output: an editable .xlsx with four
sheets (Confirmed Properties / For Review (unconfirmed) / All hostnames / Methodology & sources) and
a domains.txt of CONFIRMED registrable domains only (we do not scope on guesses).

CLI:  build_inventory.py <candidates.json> <out.xlsx> <out_domains.txt>
"""
import json
import pathlib
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DARK, AMBER, GRAY = "1E1E1E", "E8B500", "D9D9D9"
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
CHIP_FILL = {"likely": AMBER, "possible": GRAY}


def _headers(ws, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = HDR_FONT
        c.fill = PatternFill("solid", fgColor=DARK)
        c.alignment = Alignment(horizontal="left", vertical="center")


def _hyperlink(ws, row, col, url):
    if url:
        cell = ws.cell(row=row, column=col)
        cell.hyperlink = url
        cell.font = Font(color="0563C1", underline="single")  # explicit (avoid named-style dependency)


def _hosts_for(prop):
    f = prop.get("all_hosts_file")
    if f:
        try:
            return json.loads(pathlib.Path(f).read_text()).get("all_hosts", [])
        except Exception:
            pass
    return prop.get("sample_hosts") or []


def build_workbook(data):
    props = data.get("properties", []) or []
    confirmed = [p for p in props if p.get("confidence") == "confirmed"]
    review = [p for p in props if p.get("confidence") in ("likely", "possible")]
    wb = Workbook()

    ws = wb.active
    ws.title = "Confirmed Properties"
    _headers(ws, ["Domain", "Type", "Evidence", "Source", "Subdomains found", "In scope?", "Notes"])
    for p in confirmed:
        ws.append([p.get("registrable", ""), p.get("type", ""), p.get("evidence", ""),
                   p.get("source", ""), p.get("host_count", 0), "", p.get("notes", "")])
        _hyperlink(ws, ws.max_row, 4, p.get("source"))

    ws2 = wb.create_sheet("For Review (unconfirmed)")
    _headers(ws2, ["Domain", "Type", "Confidence", "Why flagged", "Source",
                   "Subdomains found", "In scope?", "Notes"])
    for p in review:
        conf = p.get("confidence", "")
        ws2.append([p.get("registrable", ""), p.get("type", ""), conf.upper(),
                    p.get("evidence", ""), p.get("source", ""), p.get("host_count", 0), "",
                    p.get("notes", "")])
        chip = ws2.cell(row=ws2.max_row, column=3)
        chip.fill = PatternFill("solid", fgColor=CHIP_FILL.get(conf, GRAY))
        chip.font = Font(bold=True, color=("FFFFFF" if conf == "likely" else "1E1E1E"))
        _hyperlink(ws2, ws2.max_row, 5, p.get("source"))

    ws3 = wb.create_sheet("All hostnames")
    _headers(ws3, ["Hostname", "Registrable domain", "Source"])
    for p in confirmed:
        for h in _hosts_for(p):
            ws3.append([h, p.get("registrable", ""), "crt.sh"])

    ws4 = wb.create_sheet("Methodology & sources")
    for row in [
        ["How this footprint was built"],
        ["Certificate Transparency (crt.sh), WHOIS registrant, SEC 10-K Exhibit 21 subsidiaries, the "
         "org's own brand/footer pages, and (when keyed) reverse-WHOIS / passive-DNS."],
        [""],
        ["Confidence definitions"],
        ["confirmed — WHOIS registrant match, SEC Exhibit-21 subsidiary, listed on the org's own "
         "brand/footer page, or a subdomain of a confirmed apex."],
        ["likely — strong web evidence (acquisition / Crunchbase / Wikipedia parent), not registrant-confirmed."],
        ["possible — shared cert / similar branding, unconfirmed."],
        [""],
        ["Owned vs observed: vendor / CDN / third-party domains a site merely loads or links to are "
         "NOT owned and are excluded below."],
    ]:
        ws4.append(row)
    for d in data.get("excluded", []) or []:
        ws4.append([f"Excluded: {d.get('domain', '')} — {d.get('why', '')}"])

    widths = {"Domain": 32, "Type": 14, "Confidence": 13, "Evidence": 48, "Why flagged": 48,
              "Source": 40, "Subdomains found": 16, "In scope?": 11, "Notes": 30,
              "Hostname": 36, "Registrable domain": 26}
    for ws_ in wb.worksheets:
        ncols = ws_.max_column or 1
        for col in range(1, ncols + 1):
            header = ws_.cell(row=1, column=col).value
            ws_.column_dimensions[get_column_letter(col)].width = widths.get(header, 80 if ncols == 1 else 28)
    return wb


def confirmed_domains(data):
    return sorted({p.get("registrable", "") for p in data.get("properties", []) or []
                   if p.get("confidence") == "confirmed" and p.get("registrable")})


def main(argv):
    if len(argv) < 3:
        sys.exit("usage: build_inventory.py <candidates.json> <out.xlsx>")
    data = json.loads(pathlib.Path(argv[1]).read_text())
    build_workbook(data).save(argv[2])
    # The .xlsx is the only file written (it carries the confirmed set on its Confirmed Properties
    # sheet). The confirmed domains are also PRINTED here, copy-pasteable into scope-calculator —
    # no separate domains.txt cluttering the deliverable folder.
    domains = confirmed_domains(data)
    print(argv[2])
    print(f"\nConfirmed domains ({len(domains)}) — ready for scope-calculator:")
    for d in domains:
        print(f"  {d}")


if __name__ == "__main__":
    main(sys.argv)
