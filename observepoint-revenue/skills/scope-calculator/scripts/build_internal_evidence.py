"""Rep-only internal evidence workbook (.xlsx) — NEVER sent to a customer.

Holds everything the customer-facing deliverables intentionally omit: page-count derivation
(census id, crawl status, raw vs defensible vs reduced), per-domain spiral/recursion notes,
assumptions-to-verify, modeled-vs-contracted, price-by-band, and the rollup-dominance flag
(spec §9.2). Enforces the sum-to-anchor invariant (spec §4.6).

Input: {customer, date?, rollup{...}, per_domain[{...}], pricing?{...}, internal?{...}}.

Key shape reference (see references/deliverables-mapping.md for full schema):
  rollup:     {spiral_adjusted_anchor, low, high, confidence, census_ids, crawl_status}
  per_domain: [{hostname, raw_urls, defensible_pages, discounted?, spiral_flag?, why?}]
  pricing:    {price_by_band?, predicted_scans, modeled_price, pricing_source?}
  internal:   {assumptions?, implied_frequency?}
"""
import json
import pathlib
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from anchor_guard import DOMINANCE_THRESHOLD, dominant_host  # noqa: F401  (re-exported; used by _derivation)

# --- ObservePoint brand authority (single source of truth) ---
sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[2] / "branding-guide" / "scripts"))
import brand_kit  # noqa: E402

FONT = brand_kit.font()["family"]
_c = brand_kit.colors()
DARK = _c["ink"].lstrip("#")
YELLOW = brand_kit.brand_yellow().lstrip("#")
LIGHT = _c["light"]["fill"].lstrip("#")
GRAY = _c["light"]["gray"].lstrip("#")
WHITE = _c["white"].lstrip("#")
RED = _c["semantic"]["alert"].lstrip("#")
_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _f(bold=False, color=DARK, size=10):
    return Font(name=FONT, bold=bold, color=color, size=size)


def _fill(hexc):
    return PatternFill("solid", fgColor=hexc)


def _check_invariant(data):
    total = sum(d["defensible_pages"] for d in data["per_domain"])
    anchor = data["rollup"]["spiral_adjusted_anchor"]
    if total != anchor:
        breakdown = ", ".join(f"{d['hostname']}={d['defensible_pages']}" for d in data["per_domain"])
        raise ValueError(f"per-domain defensible_pages sum {total} != rollup anchor {anchor} "
                         f"(per domain: {breakdown})")


def _widths(ws, widths):
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w


def _title(ws, text, row, color=DARK):
    c = ws.cell(row, 1, text)
    c.font = _f(bold=True, size=14, color=color)
    return row + 2


def _headers(ws, headers, row):
    for i, h in enumerate(headers):
        c = ws.cell(row, i + 1, h)
        c.font = _f(bold=True, color=WHITE)
        c.fill = _fill(DARK)
        c.border = _BORDER
    return row + 1


def _row(ws, row, values, *, alt=False, bold=(), fill=None):
    for i, v in enumerate(values):
        c = ws.cell(row, i + 1, v)
        c.font = _f(bold=(i in bold))
        c.border = _BORDER
        if fill:
            c.fill = _fill(fill)
        elif alt:
            c.fill = _fill(LIGHT)
        if isinstance(v, int) and not isinstance(v, bool):
            c.number_format = "#,##0"
        elif isinstance(v, float):
            c.number_format = "#,##0.##"
    return row + 1


def _derivation(wb, data):
    ws = wb.active
    ws.title = "Derivation (INTERNAL)"
    _widths(ws, [44, 26])
    r = _title(ws, "Page-count derivation — REP ONLY, do not send", 1, color=RED)
    rollup = data["rollup"]
    raw_total = sum(d["raw_urls"] for d in data["per_domain"])
    def_total = sum(d["defensible_pages"] for d in data["per_domain"])
    dom = dominant_host(data)
    pairs = [
        ("Customer", data.get("customer", "")),
        ("Date", data.get("date", "")),
        ("Census ID(s)", ", ".join(str(c) for c in rollup.get("census_ids", []))),
        ("Crawl status", rollup.get("crawl_status", "")),
        ("Confidence", rollup.get("confidence", "")),
        ("", ""),
        ("Validated pages — low / anchor / high",
         f"{rollup.get('low', '')} / {rollup.get('spiral_adjusted_anchor', '')} / {rollup.get('high', '')}"),
        ("Total raw URLs crawled", raw_total),
        ("Total defensible pages", def_total),
        ("Reduced (query-string duplicates)", raw_total - def_total),
    ]
    if dom:
        share = round(100.0 * dom["defensible_pages"] / rollup["spiral_adjusted_anchor"], 1)
        pairs += [("", ""),
                  ("⚠ DOMINANCE FLAG", f"{dom['hostname']} = {share}% of anchor — verify it is not a "
                                       f"recursion trap before quoting (spec §9.2)")]
    for label, val in pairs:
        c1 = ws.cell(r, 1, label)
        c1.font = _f(bold=True)
        c2 = ws.cell(r, 2, val)
        c2.font = _f()
        if isinstance(val, int) and not isinstance(val, bool):
            c2.number_format = "#,##0"
        elif isinstance(val, float):
            c2.number_format = "#,##0.##"
        if "DOMINANCE" in str(label):
            c1.fill = _fill(YELLOW)
            c2.fill = _fill(YELLOW)
        r += 1


def _per_domain(wb, data):
    ws = wb.create_sheet("Per-Domain (INTERNAL)")
    _widths(ws, [40, 14, 14, 14, 36])
    r = _title(ws, "Per-domain derivation", 1)
    r = _headers(ws, ["Property (domain)", "Raw URLs", "Defensible pages", "Reduced", "Note"], r)
    for i, d in enumerate(sorted(data["per_domain"], key=lambda x: -x.get("discounted", 0))):
        note = d.get("why", "") or ("query-string duplicates removed" if d.get("spiral_flag") else "")
        r = _row(ws, r, [d["hostname"], d["raw_urls"], d["defensible_pages"],
                         d.get("discounted", 0), note], alt=(i % 2 == 1))


def _assumptions(wb, data):
    internal = data.get("internal", {})
    if not internal.get("assumptions"):
        return
    ws = wb.create_sheet("Assumptions (INTERNAL)")
    _widths(ws, [80])
    r = _title(ws, "Assumptions to verify with the customer", 1)
    for a in internal["assumptions"]:
        r = _row(ws, r, [a])
    if internal.get("implied_frequency") is not None:
        r += 1
        _row(ws, r, [f"Implied blended frequency: {internal['implied_frequency']}× "
                     f"(vs the public single-frequency calculator)."])


def _pricing(wb, data):
    pr = data.get("pricing", {})
    if not pr:
        return
    ws = wb.create_sheet("Pricing (INTERNAL)")
    _widths(ws, [24, 18])
    r = _title(ws, "Pricing — predicted scans & price", 1)
    r = _row(ws, r, ["Predicted annual page scans", pr.get("predicted_scans", "")])
    r = _row(ws, r, ["Annual price (USD)", pr.get("modeled_price", "")], alt=True)
    r += 1
    if pr.get("price_by_band"):
        r = _headers(ws, ["Band width", "Rate/scan", "Scans", "Cost"], r)
        for i, b in enumerate(pr["price_by_band"]):
            band = "tail" if b.get("band_limit") is None else b["band_limit"]
            r = _row(ws, r, [band, b["rate"], b["pages"], b["cost"]], alt=(i % 2 == 1))
    if pr.get("pricing_source"):
        r += 1
        _row(ws, r, [f"Pricing source: {pr['pricing_source']}"])


def build_workbook(data):
    _check_invariant(data)
    wb = Workbook()
    _derivation(wb, data)
    _per_domain(wb, data)
    _assumptions(wb, data)
    _pricing(wb, data)
    return wb


_DOC_REF = "see references/deliverables-mapping.md"
_REQUIRED = {"rollup": "{spiral_adjusted_anchor, …}", "per_domain": "[{hostname, raw_urls, defensible_pages}, …]"}


def _validate(data):
    if not isinstance(data, dict):
        sys.exit(f"scope-calculator: malformed internal-evidence inputs — expected a JSON object; {_DOC_REF}")
    for key, shape in _REQUIRED.items():
        if key not in data or data[key] in (None, {}, []):
            sys.exit(f"scope-calculator: missing/malformed '{key}' — expected {shape}; {_DOC_REF}")
    if "spiral_adjusted_anchor" not in data["rollup"]:
        sys.exit(f"scope-calculator: 'rollup' is missing 'spiral_adjusted_anchor'; {_DOC_REF}")


def main(argv):
    if len(argv) > 1:
        with open(argv[1]) as fh:
            raw = fh.read()
    else:
        raw = sys.stdin.read()
    out = argv[2] if len(argv) > 2 else "internal-evidence.xlsx"
    try:
        data = json.loads(raw)
    except json.JSONDecodeError as e:
        sys.exit(f"scope-calculator: internal-evidence inputs are not valid JSON ({e}); {_DOC_REF}")
    _validate(data)
    build_workbook(data).save(out)
    print(out)


if __name__ == "__main__":
    main(sys.argv)
