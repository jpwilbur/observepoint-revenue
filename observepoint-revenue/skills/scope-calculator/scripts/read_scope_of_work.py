"""Reader, integrity validator, and proposal-payload builder for the AE-edited
"Scope of Work" workbook.

Public surface
--------------
read_scope_of_work(path) -> dict
    Open the workbook, run integrity validation, return compute_scope-ready inputs.
    Raises IntegrityError on any HARD-STOP finding.

validate_scope_of_work(wb_f, wb_v, *, geom) -> (hard_stops, warnings)
    Pure validator; caller decides whether to raise.

proposal_payload_from_scope_of_work(path, *, ...) -> dict
    Read + validate + compute + assemble the build_proposal.build_proposal payload.

CLI:  python read_scope_of_work.py "<file>.xlsx" [out.json]
    Prints proposal JSON or exits non-zero with the integrity report (no traceback).
"""
from __future__ import annotations

import json
import pathlib
import re
import sys
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula

import compute_scope as cs

_DOC_REF = "see references/deliverables-mapping.md"

# ---------------------------------------------------------------------------
# Custom exception
# ---------------------------------------------------------------------------

class IntegrityError(ValueError):
    """Raised when the Scope of Work workbook fails one or more HARD-STOP checks."""


# ---------------------------------------------------------------------------
# Geometry discovery helpers
# ---------------------------------------------------------------------------

def _discover_geometry(wb_f) -> dict:
    """Discover all dynamic row boundaries from the formula-view workbook.

    Returns a dict with keys:
      scope_detail:  {first, last}          — Scope Detail data rows
      scope_of_work: {buf, total, inv, n_layers, price_total_row}
      pricing:       {total_row, n_tiers}
    """
    # ---- Scope Detail ----
    if "Scope Detail" not in wb_f.sheetnames:
        return {}   # Group A will catch this
    ws_sd = wb_f["Scope Detail"]
    # header row = 3, first data row = 4
    first = 4
    last = first
    for r in range(first, ws_sd.max_row + 1):
        a = ws_sd.cell(r, 1).value
        b = ws_sd.cell(r, 2).value
        if a is None and b is None:
            break
        last = r

    # ---- Scope of Work ----
    if "Scope of Work" not in wb_f.sheetnames:
        return {"scope_detail": {"first": first, "last": last}}
    ws_sow = wb_f["Scope of Work"]
    buf = None
    for r in range(14, ws_sow.max_row + 1):
        val = ws_sow.cell(r, 1).value
        if isinstance(val, str) and val.strip() == "Buffer %":
            buf = r
            break
    if buf is None:
        return {
            "scope_detail": {"first": first, "last": last},
            "scope_of_work": {"buf": None, "total": None, "inv": None,
                               "n_layers": None, "price_total_row": None},
        }
    n_layers = buf - 14
    total = buf + 1
    inv = total + 2

    # ---- Pricing ----
    if "Pricing" not in wb_f.sheetnames:
        return {
            "scope_detail": {"first": first, "last": last},
            "scope_of_work": {"buf": buf, "total": total, "inv": inv,
                               "n_layers": n_layers, "price_total_row": None},
        }
    ws_pr = wb_f["Pricing"]
    n_tiers = None
    for r in range(5, ws_pr.max_row + 1):
        val = ws_pr.cell(r, 1).value
        if isinstance(val, str) and val.lower().startswith("recommended investment"):
            n_tiers = r - 5
            break
    price_total_row = (5 + n_tiers) if n_tiers is not None else None

    return {
        "scope_detail": {"first": first, "last": last},
        "scope_of_work": {"buf": buf, "total": total, "inv": inv,
                           "n_layers": n_layers, "price_total_row": price_total_row},
        "pricing": {"total_row": price_total_row, "n_tiers": n_tiers},
    }


# ---------------------------------------------------------------------------
# Formula normalization (for formula-integrity checks)
# ---------------------------------------------------------------------------

def _norm(formula_or_af) -> str:
    """Normalize a formula string for comparison.
    Strips leading =, uppercases, removes spaces and $, normalises sheet-quote style.
    Handles ArrayFormula objects."""
    if formula_or_af is None:
        return ""
    if isinstance(formula_or_af, ArrayFormula):
        s = formula_or_af.text or ""
    else:
        s = str(formula_or_af)
    s = s.strip()
    if s.startswith("="):
        s = s[1:]
    s = s.upper()
    s = s.replace(" ", "")
    s = s.replace("$", "")
    # Normalize sheet reference quoting: 'SCOPE DETAIL'! → SCOPEDETAIL!
    # both quoted and unquoted forms normalise to the same string
    s = re.sub(r"'([^']+)'!", lambda m: m.group(1).upper().replace(" ", "") + "!", s)
    return s


def _norm_expected(formula: str) -> str:
    return _norm(formula)


# ---------------------------------------------------------------------------
# Bool coercion
# ---------------------------------------------------------------------------

def _coerce_bool(val) -> Optional[bool]:
    """Return True/False or None if unrecognised."""
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return bool(val)
    if isinstance(val, str):
        low = val.strip().lower()
        if low in ("true", "yes", "1"):
            return True
        if low in ("false", "no", "0"):
            return False
    return None


# ---------------------------------------------------------------------------
# Integrity validator
# ---------------------------------------------------------------------------

SCOPE_DETAIL_HEADERS = [
    "Property (domain)", "Pages", "% of total",
    "Include in scope?", "Sample Size", "Notes",
]
SOW_CADENCE_HEADERS = [
    "Recommended Monitor Layer", "Recommended Cadence", "Why",
    "% of combined pages", "Runs/yr", "Pages each run", "Scans/yr",
]
PRICING_HEADERS = ["Band", "From (scans)", "To (scans)", "Rate / scan", "Cost"]


def validate_scope_of_work(wb_f, wb_v, *, geom: dict) -> tuple[list[str], list[str]]:
    """Validate the workbook.  Returns (hard_stops, warnings).

    Groups:
      A  — required sheets present
      B  — structure / headers / required rows
      C  — formula cells untampered (C1-C8 hard-stop; C9 warning)
      D  — input types & ranges (D1-D9)
      E  — reconciliation (all warnings)
    """
    hard_stops: list[str] = []
    warnings: list[str] = []

    def stop(msg: str):
        hard_stops.append(msg)

    def warn(msg: str):
        warnings.append(msg)

    # ---- Group A: required sheets ----
    for sheet in ("Scope Detail", "Scope of Work", "Pricing"):
        if sheet not in wb_f.sheetnames:
            stop(f"missing required sheet '{sheet}' — the workbook may be from an older "
                 f"template or the tab was renamed/deleted.")

    if hard_stops:
        return hard_stops, warnings   # can't proceed without sheets

    ws_sd = wb_f["Scope Detail"]
    ws_sow = wb_f["Scope of Work"]
    ws_pr = wb_f["Pricing"]
    ws_sd_v = wb_v["Scope Detail"]
    ws_sow_v = wb_v["Scope of Work"]
    ws_pr_v = wb_v["Pricing"]

    sd = geom.get("scope_detail", {})
    sow = geom.get("scope_of_work", {})
    pr = geom.get("pricing", {})
    f = sd.get("first", 4)
    l = sd.get("last", 4)
    buf = sow.get("buf")
    total = sow.get("total")
    inv = sow.get("inv")
    price_total_row = pr.get("total_row") if pr else sow.get("price_total_row")

    # ---- Group B: structure / headers / required rows ----

    # B1: Scope Detail headers
    actual_hdr = [ws_sd.cell(3, i + 1).value for i in range(6)]
    if actual_hdr != SCOPE_DETAIL_HEADERS:
        stop(f"Scope Detail headers changed (expected {SCOPE_DETAIL_HEADERS!r} in row 3, "
             f"found {actual_hdr!r}) — columns may have been inserted/deleted; the reader "
             f"can no longer locate the levers.")

    # B2: at least 1 data row
    if l < f or ws_sd.cell(f, 1).value is None:
        stop("Scope Detail has no domain rows — every property row was deleted.")

    # B3: Scope of Work structural labels
    a6 = ws_sow.cell(6, 1).value
    if not (isinstance(a6, str) and a6.strip().lower().startswith("total pages found")):
        stop(f"Scope of Work structure changed (expected 'Total Pages Found' at A6) — "
             f"a required row was deleted or moved; cannot map inputs.")

    a10 = ws_sow.cell(10, 1).value
    if not (isinstance(a10, str) and a10.strip().lower().startswith("combined")):
        stop(f"Scope of Work structure changed (expected 'Combined' at A10) — "
             f"a required row was deleted or moved; cannot map inputs.")

    a12 = ws_sow.cell(12, 1).value
    if not (isinstance(a12, str) and a12.strip().upper() == "MONITORING CADENCE"):
        stop(f"Scope of Work structure changed (expected 'MONITORING CADENCE' at A12) — "
             f"a required row was deleted or moved; cannot map inputs.")

    # Cadence header row 13
    actual_cad_hdr = [ws_sow.cell(13, i + 1).value for i in range(7)]
    if actual_cad_hdr != SOW_CADENCE_HEADERS:
        stop(f"Scope of Work structure changed (cadence header row 13 expected "
             f"{SOW_CADENCE_HEADERS!r}, found {actual_cad_hdr!r}).")

    # B4: Buffer % row must exist
    if buf is None:
        stop("Scope of Work is missing the 'Buffer %' row — cannot locate the cadence block boundary.")

    # B5: Total predicted row must exist
    if buf is not None:
        a_total = ws_sow.cell(buf + 1, 1).value
        if not (isinstance(a_total, str) and "total annual page-scans" in a_total.lower()):
            stop(f"Scope of Work is missing the 'Total annual page-scans (predicted)' row "
                 f"(expected at row {buf + 1}).")

    # B6: At least 1 cadence layer
    if buf is not None and buf <= 14:
        stop("Scope of Work has no cadence layers — every monitoring row was deleted.")

    # B7: Pricing headers and total row
    actual_pr_hdr = [ws_pr.cell(4, i + 1).value for i in range(5)]
    if actual_pr_hdr != PRICING_HEADERS:
        stop(f"Pricing headers changed (expected {PRICING_HEADERS!r} in row 4, "
             f"found {actual_pr_hdr!r}).")

    if price_total_row is None:
        stop("Pricing is missing the 'Recommended investment / year' total row.")

    # If structural checks fail here, skip formula checks that depend on geometry
    if hard_stops:
        return hard_stops, warnings

    n_layers = buf - 14

    # ---- Group C: formula integrity ----

    # C1: B6 SUMPRODUCT (ArrayFormula)
    b6_val = ws_sow["B6"].value
    expected_c1 = f"SUMPRODUCT(--'SCOPE DETAIL'!D{f}:D{l},'SCOPE DETAIL'!B{f}:B{l},'SCOPE DETAIL'!E{f}:E{l})"
    expected_c1_norm = _norm_expected(expected_c1)
    b6_norm = _norm(b6_val)
    if expected_c1_norm not in b6_norm and b6_norm != expected_c1_norm:
        # Also allow without -- prefix (some versions drop it)
        alt = _norm_expected(f"SUMPRODUCT('SCOPE DETAIL'!D{f}:D{l}*('SCOPE DETAIL'!B{f}:B{l})*('SCOPE DETAIL'!E{f}:E{l}))")
        if b6_norm != alt:
            stop(f"Total Pages Found (Scope of Work!B6) is no longer the SUMPRODUCT formula — "
                 f"it was overwritten with a literal or edited; the page total can no longer be trusted. "
                 f"(found: {b6_val!r})")

    # C2: B10 = =B6*B7*B8*B9
    b10_val = ws_sow["B10"].value
    if _norm(b10_val) != _norm_expected("B6*B7*B8*B9"):
        stop(f"Combined Page Total (B10) formula was overwritten — "
             f"expected =B6*B7*B8*B9, found {b10_val!r}.")

    # C3: Per-layer cadence formulas F{r}==$B$10*D{r}, G{r}==ROUND(F{r}*E{r},2)
    for r in range(14, buf):
        f_cell = ws_sow.cell(r, 6).value
        g_cell = ws_sow.cell(r, 7).value
        exp_f = _norm_expected(f"B10*D{r}")
        exp_g = _norm_expected(f"ROUND(F{r}*E{r},2)")
        layer_name = ws_sow.cell(r, 1).value or f"row {r}"
        if _norm(f_cell) != exp_f:
            stop(f"Cadence row {r} ('{layer_name}') lost its 'Pages each run' formula "
                 f"(expected =$B$10*D{r}, found {f_cell!r}).")
        if _norm(g_cell) != exp_g:
            stop(f"Cadence row {r} ('{layer_name}') lost its 'Scans/yr' formula "
                 f"(expected =ROUND(F{r}*E{r},2), found {g_cell!r}).")

    # C4: Buffer row formulas
    f_buf = ws_sow.cell(buf, 6).value
    g_buf = ws_sow.cell(buf, 7).value
    if _norm(f_buf) != _norm_expected(f"B10*D{buf}"):
        stop(f"Buffer row formulas (F{buf}) were overwritten — "
             f"expected =$B$10*D{buf}, found {f_buf!r}.")
    if _norm(g_buf) != _norm_expected(f"F{buf}"):
        stop(f"Buffer row formulas (G{buf}) were overwritten — "
             f"expected =F{buf}, found {g_buf!r}.")

    # C5: Predicted total G{total} = ROUND(SUM(G14:G{buf}),0)
    g_total = ws_sow.cell(total, 7).value
    exp_gtotal = _norm_expected(f"ROUND(SUM(G14:G{buf}),0)")
    if _norm(g_total) != exp_gtotal:
        stop(f"Predicted total (G{total}) is no longer the SUM of the cadence + buffer scans — "
             f"expected =ROUND(SUM(G14:G{buf}),0), found {g_total!r}.")

    # C6: Investment cell B{inv} = ='Pricing'!E{price_total_row}
    b_inv = ws_sow.cell(inv, 2).value
    exp_inv = _norm_expected(f"PRICING!E{price_total_row}")
    if _norm(b_inv) != exp_inv:
        stop(f"Investment cell (B{inv}) no longer links to the Pricing total — "
             f"expected ='Pricing'!E{price_total_row}, found {b_inv!r}.")

    # C7: Pricing cost column per band
    n_tiers = price_total_row - 5
    for j in range(n_tiers):
        n = 5 + j
        e_cell = ws_pr.cell(n, 5).value
        exp_e = _norm_expected(
            f"MAX(0,MIN('SCOPE OF WORK'!G{total},C{n})-B{n})*D{n}"
        )
        if _norm(e_cell) != exp_e:
            stop(f"Pricing band {j + 1} cost formula (E{n}) was overwritten — "
                 f"expected =MAX(0, MIN('Scope of Work'!$G${total}, C{n}) - B{n}) * D{n}, "
                 f"found {e_cell!r}.")

    # C8: Pricing total
    e_total = ws_pr.cell(price_total_row, 5).value
    exp_etotal = _norm_expected(f"ROUND(SUM(E5:E{price_total_row - 1}),2)")
    if _norm(e_total) != exp_etotal:
        stop(f"Pricing total (E{price_total_row}) is no longer the SUM of the band costs — "
             f"expected =ROUND(SUM(E5:E{price_total_row - 1}),2), found {e_total!r}.")

    # C9: % of total column in Scope Detail (WARNING only)
    for r in range(f, l + 1):
        c_cell = ws_sd.cell(r, 3).value
        exp_c = _norm_expected(f"B{r}/SUM(B{f}:B{l})")
        if _norm(c_cell) != exp_c:
            warn(f"Scope Detail row {r} '% of total' formula was overwritten "
                 f"(display-only; not used in calculations).")

    # ---- Group D: input type/range validation ----

    # Read per-domain rows for D1-D3
    per_domain_raw = []
    for r in range(f, l + 1):
        host = ws_sd.cell(r, 1).value
        pages_raw = ws_sd.cell(r, 2).value
        include_raw = ws_sd.cell(r, 4).value
        sample_raw = ws_sd.cell(r, 5).value

        # D1: pages numeric and ≥ 0
        if pages_raw is None or not isinstance(pages_raw, (int, float)):
            stop(f"Scope Detail row {r} ('{host}') has a non-numeric Pages value {pages_raw!r}.")
            pages_v = 0
        elif pages_raw < 0:
            stop(f"Scope Detail row {r} ('{host}') has a negative Pages value {pages_raw!r}.")
            pages_v = 0
        else:
            pages_v = int(pages_raw)

        # D2: Include? boolean
        inc_bool = _coerce_bool(include_raw)
        if inc_bool is None:
            stop(f"Scope Detail row {r} Include? must be TRUE or FALSE, found {include_raw!r}.")
            inc_bool = True

        # D3: Sample Size — normalization
        sample_v, sample_warn = _normalize_sample(r, host, sample_raw)
        if sample_warn and sample_warn.startswith("STOP:"):
            stop(sample_warn[5:])
            sample_v = 1.0
        elif sample_warn:
            warn(sample_warn)

        is_agg = bool(isinstance(host, str) and re.match(r"^\(\d+ additional domains", host))
        per_domain_raw.append({
            "hostname": host,
            "defensible_pages": pages_v,
            "include": inc_bool,
            "sample_size": sample_v,
            "_aggregate": is_agg,
        })

    # D4: Multipliers ≥ 1
    for cell_ref, label in (("B7", "Geographies"), ("B8", "Consent scenarios"), ("B9", "Environments")):
        val = ws_sow[cell_ref].value
        if not isinstance(val, (int, float)):
            stop(f"{label} ({cell_ref}) must be a number ≥ 1, found {val!r}.")
        elif val < 1:
            stop(f"{label} ({cell_ref}) must be a number ≥ 1, found {val!r}.")

    # D5-D7: Cadence input cells
    for r in range(14, buf):
        layer_name = ws_sow.cell(r, 1).value or f"row {r}"
        pct_val = ws_sow.cell(r, 4).value
        runs_val = ws_sow.cell(r, 5).value

        if not isinstance(pct_val, (int, float)):
            stop(f"Cadence row {r} '{layer_name}' % of combined pages must be ≥ 0, found {pct_val!r}.")
        elif pct_val < 0:
            stop(f"Cadence row {r} '{layer_name}' % of combined pages must be ≥ 0, found {pct_val!r}.")

        if not isinstance(runs_val, (int, float)):
            stop(f"Cadence row {r} '{layer_name}' Runs/yr must be a non-negative number, found {runs_val!r}.")
        elif runs_val < 0:
            stop(f"Cadence row {r} '{layer_name}' Runs/yr must be a non-negative number, found {runs_val!r}.")
        elif isinstance(runs_val, float) and runs_val != int(runs_val):
            warn(f"Cadence row {r} '{layer_name}' Runs/yr is non-integer ({runs_val}) — "
                 f"using as-is; the workbook formula rounds to 2dp.")

    # D7: Buffer %
    buf_pct_val = ws_sow.cell(buf, 4).value
    if not isinstance(buf_pct_val, (int, float)):
        stop(f"Buffer % (D{buf}) must be ≥ 0, found {buf_pct_val!r}.")
    elif buf_pct_val < 0:
        stop(f"Buffer % (D{buf}) must be ≥ 0, found {buf_pct_val!r}.")

    # D8: Pricing band from/to/rate
    for j in range(price_total_row - 5):
        n = 5 + j
        from_v = ws_pr.cell(n, 2).value
        to_v = ws_pr.cell(n, 3).value
        rate_v = ws_pr.cell(n, 4).value
        ok = True
        if not isinstance(from_v, (int, float)) or not isinstance(to_v, (int, float)):
            ok = False
        elif to_v < from_v:
            ok = False
        if not isinstance(rate_v, (int, float)) or rate_v < 0:
            ok = False
        if not ok:
            stop(f"Pricing band {j + 1} has an invalid From/To/Rate "
                 f"(From={from_v!r}, To={to_v!r}, Rate={rate_v!r}).")

    # D9: At least one in-scope domain with pages > 0 (no hard_stops yet from D1-D8 might still catch it)
    tpf = cs.total_pages_found(per_domain_raw)
    if tpf == 0:
        stop("No in-scope pages — every domain is excluded or sampled to 0%; "
             "there is nothing to price.")

    # ---- Group E: Reconciliation (warnings only) ----
    # Check if cached values exist
    b6_v = ws_sow_v["B6"].value
    b10_v = ws_sow_v["B10"].value
    g_total_v = ws_sow_v.cell(total, 7).value if total else None
    e_total_v = ws_pr_v.cell(price_total_row, 5).value if price_total_row else None

    all_none = all(x is None for x in (b6_v, b10_v, g_total_v, e_total_v))
    if all_none:
        warn("the workbook was not recalculated before saving, so cached totals are absent — "
             "skipping the reconciliation cross-check. "
             "(Open it in Excel and save to enable this check.)")
    else:
        # E1: Total Pages Found
        tol_tpf = max(1, round(0.005 * tpf))
        if b6_v is not None and isinstance(b6_v, (int, float)):
            if abs(b6_v - tpf) > tol_tpf:
                warn(f"Total Pages Found cache (B6={b6_v}) disagrees with recomputed {tpf} by "
                     f"{abs(b6_v - tpf)} — the file may not have been recalculated; "
                     f"the proposal uses the recomputed value.")

        # E2: Combined
        mult_g = ws_sow["B7"].value or 1
        mult_s = ws_sow["B8"].value or 1
        mult_e = ws_sow["B9"].value or 1
        combined_r = tpf * mult_g * mult_s * mult_e
        if b10_v is not None and isinstance(b10_v, (int, float)):
            tol_c = max(1, round(0.005 * combined_r))
            if abs(b10_v - combined_r) > tol_c:
                warn(f"Combined Page Total cache (B10={b10_v}) disagrees with recomputed "
                     f"{combined_r} — the file may not have been recalculated; "
                     f"the proposal uses the recomputed value.")

        # E3: Full engine reconciliation
        if not hard_stops:  # only if D checks passed
            try:
                layers_read = []
                for r in range(14, buf):
                    layers_read.append({
                        "name": ws_sow.cell(r, 1).value,
                        "pct": ws_sow.cell(r, 4).value,
                        "runs_per_year": ws_sow.cell(r, 5).value,
                        "why": ws_sow.cell(r, 3).value or "",
                    })
                buf_pct = ws_sow.cell(buf, 4).value or 0.0
                # Reconstruct tiers
                tiers_r = []
                lo_r = 0
                for j in range(price_total_row - 5):
                    n = 5 + j
                    from_v = ws_pr.cell(n, 2).value or 0
                    to_v = ws_pr.cell(n, 3).value or 0
                    rate_v = ws_pr.cell(n, 4).value or 0
                    tiers_r.append({"limit": max(0, int(to_v - from_v)),
                                    "pricePerPage": rate_v})
                    lo_r = int(to_v)
                # Use BAKED_TIERS if reconstruction looks off
                if any(t["limit"] <= 0 and i < price_total_row - 6 for i, t in enumerate(tiers_r)):
                    tiers_r = cs.BAKED_TIERS

                inputs_e3 = {
                    "page_count": {"low": tpf, "anchor": tpf, "high": tpf},
                    "multipliers": {"geographies": mult_g, "scenarios": mult_s, "environments": mult_e},
                    "cadence_layers": layers_read,
                    "buffer_pct": buf_pct,
                    "tiers": tiers_r,
                }
                res_e3 = cs.compute(inputs_e3)
                a_e3 = res_e3["anchor"]

                if g_total_v is not None and isinstance(g_total_v, (int, float)):
                    tol_s = max(1, round(0.005 * a_e3["predicted_scans"]))
                    if abs(g_total_v - a_e3["predicted_scans"]) > tol_s:
                        warn(f"Predicted scans cache (G{total}={g_total_v}) disagrees with "
                             f"engine recomputation ({a_e3['predicted_scans']}) — "
                             f"building the proposal from the engine recomputation.")

                if e_total_v is not None and isinstance(e_total_v, (int, float)):
                    tol_p = max(1.0, 0.005 * a_e3["price"]["total"])
                    if abs(e_total_v - a_e3["price"]["total"]) > tol_p:
                        warn(f"Price cache (E{price_total_row}={e_total_v}) disagrees with "
                             f"engine recomputation (${a_e3['price']['total']}) — "
                             f"building the proposal from the engine recomputation.")
            except Exception as exc:
                warn(f"Reconciliation check failed unexpectedly: {exc}")

    return hard_stops, warnings


def _normalize_sample(row: int, host, raw) -> tuple[float, str]:
    """Normalize a Sample Size cell value.
    Returns (normalized_float, warning_message_or_empty).
    If HARD-STOP needed, returns (1.0, 'STOP: <message>').
    """
    if raw is None:
        return 1.0, ""
    if not isinstance(raw, (int, float)):
        return 1.0, f"STOP: Scope Detail row {row} ('{host}') Sample Size must be a percentage between 0% and 100%, found {raw!r}."

    s = float(raw)
    if s <= 0 or s > 100:
        return 1.0, f"STOP: Scope Detail row {row} ('{host}') Sample Size must be a percentage between 0% and 100%, found {raw!r}."

    if 1 < s <= 100:
        # Entered as a bare percent (e.g. 50 meaning 50%)
        normalized = s / 100.0
        msg = (f"Scope Detail row {row} ('{host}') Sample Size read as {s:.0f}% "
               f"(entered as {raw}); interpreting as {normalized}.")
        return normalized, msg

    # s is in (0, 1] — correct
    return s, ""


# ---------------------------------------------------------------------------
# Reader
# ---------------------------------------------------------------------------

def read_scope_of_work(path: str) -> dict:
    """Open the workbook at *path*, validate integrity, and return compute_scope-ready inputs.

    Raises IntegrityError if any HARD-STOP finding is found.
    Warnings are attached as result['_warnings'] and printed to stderr.
    """
    wb_f = load_workbook(path, data_only=False)
    wb_v = load_workbook(path, data_only=True)

    geom = _discover_geometry(wb_f)
    hard_stops, warnings = validate_scope_of_work(wb_f, wb_v, geom=geom)

    if hard_stops:
        bullet_list = "\n".join(f"  • {s}" for s in hard_stops)
        fname = pathlib.Path(path).name
        raise IntegrityError(
            f"scope-calculator: the edited Scope of Work failed integrity checks — nothing was built.\n"
            f"{bullet_list}\n"
            f"Fix these in '{fname}' (or rebuild it from scratch) and try again; {_DOC_REF}"
        )

    if warnings:
        for w in warnings:
            print(f"scope-calculator [WARNING]: {w}", file=sys.stderr)

    # Extract inputs from the formula view
    ws_sd = wb_f["Scope Detail"]
    ws_sow = wb_f["Scope of Work"]
    ws_pr = wb_f["Pricing"]

    sd = geom["scope_detail"]
    sow_g = geom["scope_of_work"]
    pr_g = geom.get("pricing", {})
    f = sd["first"]
    l = sd["last"]
    buf = sow_g["buf"]
    total = sow_g["total"]
    inv = sow_g["inv"]
    price_total_row = pr_g.get("total_row") or sow_g.get("price_total_row")

    # per_domain
    per_domain = []
    for r in range(f, l + 1):
        host = ws_sd.cell(r, 1).value
        pages_raw = ws_sd.cell(r, 2).value
        include_raw = ws_sd.cell(r, 4).value
        sample_raw = ws_sd.cell(r, 5).value

        pages_v = int(pages_raw) if isinstance(pages_raw, (int, float)) and pages_raw >= 0 else 0
        inc_bool = _coerce_bool(include_raw)
        if inc_bool is None:
            inc_bool = True
        sample_v, _ = _normalize_sample(r, host, sample_raw)
        is_agg = bool(isinstance(host, str) and re.match(r"^\(\d+ additional domains", host))

        per_domain.append({
            "hostname": host,
            "defensible_pages": pages_v,
            "include": inc_bool,
            "sample_size": sample_v,
            "_aggregate": is_agg,
        })

    # multipliers
    mult_g = ws_sow["B7"].value or 1
    mult_s = ws_sow["B8"].value or 1
    mult_e = ws_sow["B9"].value or 1
    multipliers = {"geographies": mult_g, "scenarios": mult_s, "environments": mult_e}

    # cadence layers
    cadence_layers = []
    for r in range(14, buf):
        cadence_layers.append({
            "name": ws_sow.cell(r, 1).value or "",
            "why": ws_sow.cell(r, 3).value or "",
            "pct": float(ws_sow.cell(r, 4).value or 0),
            "runs_per_year": ws_sow.cell(r, 5).value or 0,
        })

    buffer_pct = float(ws_sow.cell(buf, 4).value or 0)

    # tiers: reconstruct from Pricing sheet
    tiers = []
    lo = 0
    for j in range(price_total_row - 5):
        n = 5 + j
        from_v = ws_pr.cell(n, 2).value or 0
        to_v = ws_pr.cell(n, 3).value or 0
        rate_v = ws_pr.cell(n, 4).value or 0
        # Last band: to_v may be the 10**12 sentinel — keep it as limit width
        limit = max(0, int(to_v - from_v))
        tiers.append({"limit": limit, "pricePerPage": float(rate_v)})
        lo = int(to_v)

    # Authoritative Total Pages Found via compute_scope
    tpf = cs.total_pages_found(per_domain)

    # Customer name from A3 of Scope of Work
    a3 = ws_sow["A3"].value or ""
    if isinstance(a3, str) and a3.lower().startswith("prepared for "):
        customer = a3[len("Prepared for "):].strip()
    else:
        customer = a3.strip() or pathlib.Path(path).stem.replace(" - Scope of Work", "").strip()

    # Cached values from value view
    ws_sow_v = wb_v["Scope of Work"]
    ws_pr_v = wb_v["Pricing"]
    _cached = {
        "total_pages_found": ws_sow_v["B6"].value,
        "combined": ws_sow_v["B10"].value,
        "predicted_scans": ws_sow_v.cell(total, 7).value if total else None,
        "price_total": ws_pr_v.cell(price_total_row, 5).value if price_total_row else None,
    }

    return {
        "customer": customer,
        "page_count": {"low": tpf, "anchor": tpf, "high": tpf},
        "multipliers": multipliers,
        "cadence_layers": cadence_layers,
        "buffer_pct": buffer_pct,
        "per_domain": per_domain,
        "tiers": tiers,
        "_cached": _cached,
        "_warnings": warnings,
    }


# ---------------------------------------------------------------------------
# Proposal payload builder
# ---------------------------------------------------------------------------

def proposal_payload_from_scope_of_work(
    path: str,
    *,
    prepared_by: Optional[str] = None,
    date: Optional[str] = None,
    use_case: Optional[str] = None,
    regulations: Optional[list] = None,
    monitoring_summary: str,
    properties_note: Optional[str] = None,
    consent_names: Optional[list] = None,
) -> dict:
    """Read + validate + compute → assemble the build_proposal.build_proposal payload.

    Raises IntegrityError if the workbook fails validation.
    """
    r = read_scope_of_work(path)

    # Authoritative recompute via engine
    inputs = {
        "page_count": r["page_count"],
        "multipliers": r["multipliers"],
        "cadence_layers": r["cadence_layers"],
        "buffer_pct": r["buffer_pct"],
        "tiers": r["tiers"],
        "customer": r["customer"],
    }
    res = cs.compute(inputs)
    anchor = res["anchor"]

    # Domains = all in-scope hostnames (exclude aggregate label; it's not a URL)
    domains = [
        d["hostname"] for d in r["per_domain"]
        if d.get("include", True) and not d.get("_aggregate", False)
    ]

    names = consent_names or ["Default"]

    # CRITICAL CORRECTION: emit the keys build_proposal.build_proposal requires:
    #   usage = {combined_pages, predicted_scans}
    #   pricing = {predicted_price, price_by_band, ...}
    #   multipliers verbatim from engine output
    #   cadence_layers from anchor["cadence_by_layer"]
    payload = {
        "customer": r["customer"],
        "prepared_by": prepared_by,
        "date": date,
        "use_case": use_case,
        "regulations": regulations or [],
        "domains": domains,
        "properties_note": properties_note,
        "monitoring_summary": monitoring_summary,
        "page_count": r["page_count"],   # low=anchor=high (confirmed scope)
        "consent_states": {
            "count": r["multipliers"]["scenarios"],
            "names": names,
        },
        # Engine-emitted multipliers (verbatim) — required for _sweep_reconcile_error
        "multipliers": {
            "geographies": res["multipliers"]["geographies"],
            "scenarios": res["multipliers"]["scenarios"],
            "environments": res["multipliers"]["environments"],
        },
        # Engine cadence_by_layer carries pct + why + pages + runs (build_proposal re-derives from pct)
        "cadence_layers": anchor["cadence_by_layer"],
        "buffer_pct": r["buffer_pct"],
        # usage: required by build_proposal
        "usage": {
            "combined_pages": anchor["combined_pages"],
            "predicted_scans": anchor["predicted_scans"],
        },
        # pricing: predicted_price is the key build_proposal requires
        "pricing": {
            "predicted_price": anchor["price"]["total"],
            "price_by_band": anchor["price"]["breakdown"],
            "pricing_source": res.get("pricing_source", f"baked ({cs.BAKED_AS_OF})"),
            # Omit range_low_price / range_high_price for confirmed single scope
            # (low==anchor==high, so range paragraph would render "$X–$X" — not useful)
        },
    }

    return payload


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _main(argv):
    if len(argv) < 2:
        print("Usage: read_scope_of_work.py <Scope of Work.xlsx> [out.json]", file=sys.stderr)
        sys.exit(1)

    path = argv[1]
    out_path = argv[2] if len(argv) > 2 else None

    try:
        payload = proposal_payload_from_scope_of_work(
            path,
            monitoring_summary="[no monitoring summary provided — pass via API]",
        )
    except IntegrityError as exc:
        print(str(exc), file=sys.stderr)
        sys.exit(1)
    except FileNotFoundError:
        print(f"scope-calculator: file not found: {path}", file=sys.stderr)
        sys.exit(1)
    except Exception as exc:
        print(f"scope-calculator: unexpected error: {exc}", file=sys.stderr)
        sys.exit(1)

    result_json = json.dumps(payload, indent=2, default=str)
    if out_path:
        pathlib.Path(out_path).write_text(result_json)
        print(out_path)
    else:
        print(result_json)


if __name__ == "__main__":
    _main(sys.argv)
