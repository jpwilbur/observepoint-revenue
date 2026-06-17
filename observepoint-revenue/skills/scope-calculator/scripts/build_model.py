"""Live Excel **Scope of Work** workbook (.xlsx) for the scope-calculator.

Customer-facing workbook driven by the Scope Detail tab: Total Pages Found is a SUMPRODUCT over
in-scope domains × sample size; Combined = × geographies × consent × environments; a 4-(or N-)layer
priority cadence plus an additive Buffer row; price = graduated tiers on the predicted total.
Yellow cells mark the editable levers; sheets are NOT protected (integrity is validated at
proposal-generation time instead). The formulas reproduce compute_scope.py exactly on recalc, and
a dependency-free emulator asserts that in the tests.

Sheets (in order): Scope Detail · Scope of Work · Pricing · Sample pages.
"""
import pathlib
import re
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

import customer_clean

# ---------- theme ----------
FONT = "Montserrat"
DARK, YELLOW, LIGHT, GRAY, WHITE = "1E1E1E", "F2CD14", "F2F2F2", "5C5C5C", "FFFFFF"
INPUT_FILL = "FFF7CC"   # pale yellow — marks the editable levers
LOGO = pathlib.Path(__file__).resolve().parent.parent / "assets" / "op-logo.png"

_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_YBAR = Side(style="thick", color=YELLOW)

TOP_N = 20   # individual domains listed; the rest collapse into one bottom aggregate row
# Stage 1 may already hand us a long-tail aggregate row (the census itemizes only ~top-40); detect it
# so we fold it INTO the single bottom aggregate instead of ranking it as if it were one domain.
_AGG_RE = re.compile(r"^\(\s*(\d+)\s+additional domains")

# Runs/yr → customer-facing cadence word.
_CADENCE_WORD = {1: "Yearly", 4: "Quarterly", 12: "Monthly", 26: "Bi-weekly", 52: "Weekly", 365: "Daily"}


def _cadence_word(runs):
    return _CADENCE_WORD.get(runs, f"{runs}×/yr")


def _f(bold=False, color=DARK, size=10):
    return Font(name=FONT, bold=bold, color=color, size=size)


def _fill(hexc):
    return PatternFill("solid", fgColor=hexc)


def _widths(ws, widths):
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w


def _title(ws, text, span, row):
    c = ws.cell(row, 1, text)
    c.font = _f(bold=True, size=15)
    for i in range(span):
        ws.cell(row, i + 1).border = Border(bottom=_YBAR)
    ws.row_dimensions[row].height = 22
    return row + 2


def _headers(ws, headers, row):
    for i, h in enumerate(headers):
        c = ws.cell(row, i + 1, h)
        c.font = _f(bold=True, color=WHITE)
        c.fill = _fill(DARK)
        c.border = _BORDER
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 26
    return row + 1


def _lever(cell, fmt):
    """Style a cell as an editable yellow lever (no protection — sheets are unprotected)."""
    cell.fill = _fill(INPUT_FILL)
    cell.number_format = fmt
    cell.font = _f()


# ---------- Scope Detail (first sheet) ----------

def _scope_detail(wb, data):
    """Build the Scope Detail sheet: top-20 domains individually, the rest collapsed into one
    bottom aggregate row. NOTE: that aggregate row is a SINGLE Include?/Sample Size lever governing
    the whole long tail as a block — toggling it does not preserve the per-domain engine breakdown."""
    ws = wb.active
    ws.title = "Scope Detail"
    _widths(ws, [44, 14, 12, 16, 12, 30])
    r = _title(ws, "Scope detail", 6, 1)
    hdr_row = r
    r = _headers(ws, ["Property (domain)", "Pages", "% of total",
                      "Include in scope?", "Sample Size", "Notes"], r)
    ws.freeze_panes = ws.cell(r, 1)

    # Separate any pre-existing long-tail aggregate row(s) from the individual domains, so a Stage-1
    # aggregate isn't ranked as a domain (it would sort to the top) or double-counted. Fold it into
    # the single bottom aggregate together with the individuals beyond the top-20.
    individuals, agg_pages, agg_domains = [], 0, 0
    for d in data["per_domain"]:
        m = _AGG_RE.match(str(d["hostname"]))
        if m:
            agg_pages += d["defensible_pages"]
            agg_domains += int(m.group(1))
        else:
            individuals.append(d)
    ordered = sorted(individuals, key=lambda x: -x["defensible_pages"])
    tail = ordered[TOP_N:]
    rows = [(d["hostname"], d["defensible_pages"]) for d in ordered[:TOP_N]]
    extra_domains = agg_domains + len(tail)
    extra_pages = agg_pages + sum(d["defensible_pages"] for d in tail)
    if extra_domains > 0:
        rows.append((f"({extra_domains} additional domains — long tail, aggregated)", extra_pages))

    first = r
    last = r + len(rows) - 1
    for i, (host, pages) in enumerate(rows):
        n = r + i
        alt = (i % 2 == 1)
        fill = LIGHT if alt else WHITE
        c = ws.cell(n, 1, host); c.font = _f(); c.fill = _fill(fill); c.border = _BORDER
        c = ws.cell(n, 2, pages); c.font = _f(); c.fill = _fill(fill); c.border = _BORDER
        c.number_format = "#,##0"
        c = ws.cell(n, 3, f"=B{n}/SUM($B${first}:$B${last})")
        c.font = _f(); c.fill = _fill(fill); c.border = _BORDER; c.number_format = "0.0%"
        c = ws.cell(n, 4, True); c.border = _BORDER; c.alignment = Alignment(horizontal="center")
        _lever(c, "General")
        c = ws.cell(n, 5, 1.0); c.border = _BORDER
        _lever(c, "0%")
        c = ws.cell(n, 6, None); c.border = _BORDER; c.fill = _fill(fill)
    return {"first": first, "last": last}


# ---------- Scope of Work tab ----------

def _scope_of_work(wb, data, detail):
    ws = wb.create_sheet("Scope of Work")
    m = data.get("multipliers", {})
    layers = data["cadence_layers"]
    f, l = detail["first"], detail["last"]

    if LOGO.exists():
        try:
            from openpyxl.drawing.image import Image as XLImage
            img = XLImage(str(LOGO)); img.width, img.height = 200, 31
            ws.add_image(img, "A1")
        except Exception:
            pass
    ws.row_dimensions[1].height = 30
    ws["A2"] = "Scope of Work"; ws["A2"].font = _f(bold=True, size=16)
    ws["A3"] = f"Prepared for {data.get('customer', '')}"; ws["A3"].font = _f(color=GRAY, size=10)

    ws["A5"] = "INPUTS"; ws["A5"].font = _f(bold=True, size=11)

    def label(row, text, note):
        ws[f"A{row}"] = text; ws[f"A{row}"].font = _f(bold=True)
        ws[f"C{row}"] = note; ws[f"C{row}"].font = _f(color=GRAY, size=9)

    label(6, "Total Pages Found", "Sum of in-scope domain pages, linked from Scope detail")
    b6 = ws["B6"]
    b6.value = ArrayFormula("B6", f"=SUMPRODUCT(--'Scope Detail'!D{f}:D{l},"
                                  f"'Scope Detail'!B{f}:B{l},'Scope Detail'!E{f}:E{l})")
    b6.number_format = "#,##0"; b6.font = _f()

    label(7, "Geographies", "State / Country / Jurisdiction")
    _lever(ws["B7"], '"× "0'); ws["B7"].value = m.get("geographies", 1)
    label(8, "Consent scenarios", "Opt-out/GPC/Default consent/etc.")
    _lever(ws["B8"], '"× "0'); ws["B8"].value = m.get("scenarios", 1)
    label(9, "Environments", "Prod / Pre-Prod / Authenticated / etc.")
    _lever(ws["B9"], '"× "0'); ws["B9"].value = m.get("environments", 1)

    ws["A10"] = "Combined Page Total"; ws["A10"].font = _f(bold=True)
    ws["B10"] = "=B6*B7*B8*B9"; ws["B10"].number_format = "#,##0"
    ws["C10"] = '=TEXT(B6,"#,##0")&" × "&B7&" × "&B8&" × "&B9&" = "&TEXT(B10,"#,##0")'
    ws["C10"].font = _f(color=GRAY, size=9)

    ws["A12"] = "MONITORING CADENCE"; ws["A12"].font = _f(bold=True, size=11)
    _headers(ws, ["Recommended Monitor Layer", "Recommended Cadence", "Why",
                  "% of combined pages", "Runs/yr", "Pages each run", "Scans/yr"], 13)

    row = 14
    for L in layers:
        ws.cell(row, 1, L["name"]).font = _f()
        ws.cell(row, 2, _cadence_word(L["runs_per_year"])).font = _f()
        c = ws.cell(row, 3, L.get("why", "")); c.font = _f(); c.alignment = Alignment(wrap_text=True)
        _lever(ws.cell(row, 4, L["pct"]), "0.##%")
        _lever(ws.cell(row, 5, L["runs_per_year"]), "#,##0")
        ws.cell(row, 6, f"=$B$10*D{row}").number_format = "#,##0"
        ws.cell(row, 7, f"=ROUND(F{row}*E{row},2)").number_format = "#,##0"
        row += 1

    buf = row
    ws.cell(buf, 1, "Buffer %").font = _f(bold=True)
    c = ws.cell(buf, 3, "Ad-hoc testing and projects regularly push scanning needs past the "
                        "scheduled monitoring."); c.font = _f(); c.alignment = Alignment(wrap_text=True)
    _lever(ws.cell(buf, 4, data.get("buffer_pct", 0.0)), "0%")
    ws.cell(buf, 6, f"=$B$10*D{buf}").number_format = "#,##0"
    ws.cell(buf, 7, f"=F{buf}").number_format = "#,##0"

    total = buf + 1
    ws.cell(total, 1, "Total annual page-scans (predicted)").font = _f(bold=True)
    g = ws.cell(total, 7, f"=ROUND(SUM(G14:G{buf}),0)"); g.font = _f(bold=True); g.number_format = "#,##0"

    import compute_scope as _cs
    tiers = data.get("tiers") or _cs.BAKED_TIERS
    price_total_row = 5 + len(tiers)
    inv = total + 2
    ws.cell(inv, 1, "Recommended investment / year (USD)").font = _f(bold=True, size=12)
    b = ws.cell(inv, 2, f"='Pricing'!E{price_total_row}")
    b.number_format = "$#,##0"; b.fill = _fill(YELLOW); b.font = _f(bold=True, size=12)

    ws.cell(inv + 2, 1, "Yellow cells are editable — change them and the totals/price "
                        "update automatically.").font = _f(color=GRAY, size=9)
    _widths(ws, [34, 16, 46, 18, 10, 14, 14])
    return {"predicted_row": total}


# ---------- Pricing ----------

def _pricing(wb, data, sow):
    import compute_scope as _cs
    tiers = data.get("tiers") or _cs.BAKED_TIERS
    pr = sow["predicted_row"]
    ws = wb.create_sheet("Pricing")
    ws["A2"] = "ObservePoint published pricing — graduated tiers"
    ws["A2"].font = _f(bold=True, size=12)
    _headers(ws, ["Band", "From (scans)", "To (scans)", "Rate / scan", "Cost"], 4)
    lo = 0
    for i, t in enumerate(tiers):
        n = 5 + i
        hi = 10 ** 12 if i == len(tiers) - 1 else lo + t["limit"]
        ws.cell(n, 1, i + 1).font = _f()
        ws.cell(n, 2, lo).number_format = "#,##0"
        ws.cell(n, 3, hi).number_format = "#,##0"
        ws.cell(n, 4, t["pricePerPage"]).number_format = "$#,##0.00"
        ws.cell(n, 5, f"=MAX(0, MIN('Scope of Work'!$G${pr}, C{n}) - B{n}) * D{n}").number_format = "$#,##0.00"
        lo += t["limit"]
    total_row = 5 + len(tiers)
    ws.cell(total_row, 1, "Recommended investment / year").font = _f(bold=True)
    ws.cell(total_row, 5, f"=ROUND(SUM(E5:E{total_row - 1}),2)").number_format = "$#,##0"
    ws.cell(total_row + 2, 1, "Tier bands and per-scan rates mirror ObservePoint's published "
                              "pricing model.").font = _f(color=GRAY, size=9)
    _widths(ws, [8, 16, 16, 14, 16])


# ---------- Sample pages ----------

def _sample_pages(wb, data):
    ws = wb.create_sheet("Sample pages")
    _widths(ws, [34, 70])
    r = _title(ws, "Sample pages — real examples found on each property", 2, 1)
    note = ws.cell(r, 1, "A handful of real example pages per property (the largest by page count) "
                         "— so you can see these are genuine pages.")
    note.font = _f(color=GRAY, size=9); note.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.row_dimensions[r].height = 26
    r = _headers(ws, ["Property (domain)", "Example page URL"], r + 1)
    ws.freeze_panes = ws.cell(r, 1)
    any_rows, alt = False, False
    for d in sorted(data["per_domain"], key=lambda x: -x["defensible_pages"]):
        for s in d.get("url_samples", []):
            fill = LIGHT if alt else WHITE
            ws.cell(r, 1, d["hostname"]).fill = _fill(fill)
            ws.cell(r, 2, s).fill = _fill(fill)
            for col in (1, 2):
                ws.cell(r, col).font = _f(); ws.cell(r, col).border = _BORDER
            r += 1; any_rows = True
        if d.get("url_samples"):
            alt = not alt
    if not any_rows:
        ws.cell(r, 1, "(no per-URL samples captured)").font = _f()


# ---------- public API ----------

def build_workbook(data):
    """Build the live Scope of Work workbook. Sheets: Scope Detail · Scope of Work · Pricing · Sample pages.
    No sheet protection — yellow cells mark the levers; integrity is validated at proposal time."""
    layers = data.get("cadence_layers", [])
    if not (1 <= len(layers) <= 6):
        raise ValueError(f"scope of work: expected 1–6 cadence layers; got {len(layers)}.")
    if not data.get("per_domain"):
        raise ValueError("scope of work: 'per_domain' is empty — nothing to scope.")
    strings = [L.get("name", "") for L in layers] + [L.get("why", "") for L in layers]
    customer_clean.assert_clean(strings, where="scope of work")

    wb = Workbook()
    detail = _scope_detail(wb, data)
    sow = _scope_of_work(wb, data, detail)
    _pricing(wb, data, sow)
    _sample_pages(wb, data)
    return wb


_DOC_REF = "see references/deliverables-mapping.md"
_REQUIRED = {
    "page_count": "{low, anchor, high}",
    "cadence_layers": "[{name, pct, runs_per_year}, …]",
    "per_domain": "[{hostname, defensible_pages}, …]",
}


def _validate(data):
    if not isinstance(data, dict):
        sys.exit(f"scope-calculator: malformed model inputs — expected a JSON object; {_DOC_REF}")
    for key, shape in _REQUIRED.items():
        if key not in data or data[key] in (None, {}, []):
            sys.exit(f"scope-calculator: missing/malformed '{key}' — expected {shape}; {_DOC_REF}")
    pc = data["page_count"]
    if not isinstance(pc, dict) or any(k not in pc for k in ("low", "anchor", "high")):
        sys.exit(f"scope-calculator: missing/malformed 'page_count' — expected "
                 f"{_REQUIRED['page_count']}; {_DOC_REF}")


def main(argv):
    import json
    if len(argv) > 1:
        with open(argv[1]) as fh:
            raw = fh.read()
    else:
        raw = sys.stdin.read()
    out = argv[2] if len(argv) > 2 else "scope-of-work.xlsx"
    try:
        data = json.loads(raw)
    except json.JSONDecodeError as e:
        sys.exit(f"scope-calculator: model inputs are not valid JSON ({e}); {_DOC_REF}")
    _validate(data)
    build_workbook(data).save(out)
    print(out)


if __name__ == "__main__":
    main(sys.argv)
