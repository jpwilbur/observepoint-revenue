"""Rank territory-discovery candidates and maintain the seen-candidates log (deterministic).

The model FINDS and CLASSIFIES candidates (territory judgment, trigger classification, evidence);
this script does the mechanics: validation, seen-log dedup, ranking by trigger points x the SAME
recency decay research-account uses (scoring-config.json is passed in by path — single source of
truth, never copied), chat-ready stdout, and an optional .xlsx radar.

The decay math is reimplemented (not imported across skills) and pinned by a parity test against
score_account.recency_factor. `date` in the candidates JSON is the as_of clock (not the system
clock) so a given file always ranks the same.

CLI:  rank_candidates.py <candidates.json> <scoring-config.json>
                         [--seen <seen.json>] [--include-seen] [--xlsx <out.xlsx>]
"""
import argparse
import json
import math
import pathlib
import re
import sys
from datetime import datetime, timezone

MONTH_MS = 30 * 86_400_000  # 30-day month (parity with score_account.py / NERD scoring.ts)


def _round_half_up(x):
    """Match JS Math.round (half rounds up), not Python's banker's rounding."""
    return int(math.floor(x + 0.5))


def _parse_ms(s):
    """Parse YYYY, YYYY-MM, YYYY-MM-DD, or full ISO to epoch ms (UTC). None if unparseable.
    Date-only formats anchored to UTC so recency is machine-independent (see score_account.py)."""
    if not s:
        return None
    s = str(s).strip()
    for fmt in ("%Y-%m-%d", "%Y-%m", "%Y"):
        try:
            dt = datetime.strptime(s[:10] if fmt == "%Y-%m-%d" else s, fmt)
            return int(dt.replace(tzinfo=timezone.utc).timestamp() * 1000)
        except ValueError:
            continue
    try:
        dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        return int(dt.timestamp() * 1000)
    except Exception:
        return None


def recency_factor(date_str, recency, now_ms):
    """Full strength within fullStrengthMonths, linear decay to a 0.1 floor by zeroStrengthMonths.
    Undated/unparseable triggers get 0.6 (matched, but we can't time them)."""
    t = _parse_ms(date_str)
    if t is None:
        return 0.6
    months = (now_ms - t) / MONTH_MS
    if months <= recency["fullStrengthMonths"]:
        return 1.0
    if months >= recency["zeroStrengthMonths"]:
        return 0.1
    span = max(1, recency["zeroStrengthMonths"] - recency["fullStrengthMonths"])
    return max(0.1, 1 - (months - recency["fullStrengthMonths"]) / span)


def normalize_name(name):
    """'The Example Health-System, Inc.' == 'example health system inc' for dedup purposes.
    Lowercases, drops a leading standalone 'the' article, then strips punctuation/whitespace.
    Unicode letters and digits are kept ('Café' stays 'café')."""
    tokens = re.findall(r"[^\W_]+", (name or "").lower())
    if tokens and tokens[0] == "the":
        tokens = tokens[1:]
    return "".join(tokens)


def load_seen(path):
    """Read the seen-log; a missing, corrupt, or wrong-shaped file is just an empty log
    (it gets recreated on the next save). State must never block a discovery run."""
    if not path:
        return {"candidates": []}
    try:
        data = json.loads(pathlib.Path(path).read_text())
        if isinstance(data, dict) and isinstance(data.get("candidates"), list):
            return data
    except Exception:
        pass
    return {"candidates": []}


def rank(data, config, seen=None, include_seen=False):
    """Validate + rank candidates. Returns (ranked, dropped_count, new_log_entries).

    Previously-seen names are dropped (or kept and annotated with firstSeen when include_seen);
    every NEW name yields a seen-log entry. Validation runs on every candidate, dropped or not.
    """
    why, recency = config["whyNow"], config["recency"]
    now_ms = _parse_ms(data.get("date"))
    if now_ms is None:
        raise ValueError("candidates JSON needs a 'date' (YYYY-MM-DD) to rank against")
    seen_by_name = {normalize_name(c.get("name")): c
                    for c in (seen or {}).get("candidates", [])}

    ranked, new_entries, dropped = [], [], 0
    logged_this_run = set()
    for c in data.get("candidates", []) or []:
        norm = normalize_name(c.get("name"))
        if not norm:
            raise ValueError(f"missing or empty candidate name: {c.get('name')!r} — "
                             "every candidate needs a real company name")
        key = c.get("triggerKey")
        if key not in why:
            raise ValueError(f"unknown triggerKey {key!r} for {c.get('name')!r} — "
                             "must be one of the scoring-config whyNow keys")
        if not (c.get("sourceUrl") or "").strip():
            raise ValueError(f"missing sourceUrl for {c.get('name')!r} — "
                             "every candidate needs a real source")
        prior = seen_by_name.get(norm)
        if prior and not include_seen:
            dropped += 1
            continue
        entry = dict(c)
        entry["triggerLabel"] = why[key]["label"]
        entry["effectivePoints"] = _round_half_up(
            why[key]["points"] * recency_factor(c.get("triggerDate"), recency, now_ms))
        if prior:
            entry["firstSeen"] = prior.get("firstSeen")
        elif norm not in logged_this_run:
            logged_this_run.add(norm)
            new_entries.append({"name": c.get("name"), "firstSeen": data.get("date"),
                                "triggerKey": key, "sourceUrl": c.get("sourceUrl")})
        ranked.append(entry)
    # points desc, then newer trigger first; undated (None→0) lands last among point-ties.
    ranked.sort(key=lambda e: (-e["effectivePoints"], -(_parse_ms(e.get("triggerDate")) or 0)))
    return ranked, dropped, new_entries


RADAR_HEADERS = ["Rank", "Company", "Vertical", "Trigger", "Trigger date", "Why now",
                 "Source", "First seen", "Pursue?", "Notes"]
RADAR_WIDTHS = {"Rank": 6, "Company": 30, "Vertical": 20, "Trigger": 36, "Trigger date": 12,
                "Why now": 60, "Source": 44, "First seen": 12, "Pursue?": 10, "Notes": 30}
DARK = "1E1E1E"


def build_radar(ranked):
    """One-sheet working list: ranked candidates + fillable Pursue?/Notes for the rep."""
    from openpyxl import Workbook  # lazy: the default chat-first path must not need openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = "Discovery radar"
    for i, h in enumerate(RADAR_HEADERS, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor=DARK)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = RADAR_WIDTHS[h]
    for i, e in enumerate(ranked, 1):
        ws.append([i, e.get("name", ""), e.get("vertical", ""), e.get("triggerLabel", ""),
                   e.get("triggerDate", ""), e.get("reason", ""), e.get("sourceUrl", ""),
                   e.get("firstSeen", ""), "", ""])
        src = ws.cell(row=ws.max_row, column=7)
        src.hyperlink = e.get("sourceUrl", "")
        src.font = Font(color="0563C1", underline="single")  # explicit (avoid named-style dependency)
    return wb


def render_chat(ranked, dropped):
    lines = [f"Ranked candidates ({len(ranked)}):"]
    for i, e in enumerate(ranked, 1):
        seen_note = f"  [seen {e['firstSeen']}]" if e.get("firstSeen") else ""
        lines.append(f"{i}. {e.get('name', '')} — {e['triggerLabel']} "
                     f"({e['effectivePoints']} pts, {e.get('triggerDate') or 'undated'})"
                     f"{seen_note}")
        lines.append(f"   {e.get('reason', '')}")
        lines.append(f"   {e.get('sourceUrl', '')}")
    if dropped:
        lines.append("")
        lines.append(f"Excluded {dropped} previously-seen candidate(s) — "
                     "rerun with --include-seen to show them.")
    return "\n".join(lines)


def main(argv=None):
    ap = argparse.ArgumentParser(
        description="Rank discovery candidates; maintain the seen-candidates log.")
    ap.add_argument("candidates", help="candidates JSON assembled by the skill")
    ap.add_argument("config", help="path to research-account's scoring-config.json")
    ap.add_argument("--seen", help="seen-candidates log (created/extended as needed)")
    ap.add_argument("--include-seen", action="store_true",
                    help="keep previously-seen names in the output (refresh mode)")
    ap.add_argument("--xlsx", help="also write the discovery-radar workbook here")
    a = ap.parse_args(argv)

    try:
        data = json.loads(pathlib.Path(a.candidates).read_text())
        config = json.loads(pathlib.Path(a.config).read_text())
    except (OSError, ValueError) as e:
        sys.exit(f"could not read inputs: {e}")
    seen = load_seen(a.seen)
    try:
        ranked, dropped, new_entries = rank(data, config, seen, a.include_seen)
    except ValueError as e:
        sys.exit(str(e))
    print(render_chat(ranked, dropped))
    if a.xlsx:
        try:
            wb = build_radar(ranked)
        except ImportError:
            sys.exit("openpyxl is required for --xlsx (pip install openpyxl)")
        out = pathlib.Path(a.xlsx)
        try:
            out.parent.mkdir(parents=True, exist_ok=True)
            wb.save(out)
        except OSError as e:
            sys.exit(f"could not write {out}: {e}")
        print(f"\n{out}")
    # State last, atomically: artifacts before log, so a failed run never strands names in
    # the log (a name that resurfaces next run is safe; a logged-but-never-shown one is not).
    if a.seen and new_entries:
        p = pathlib.Path(a.seen)
        p.parent.mkdir(parents=True, exist_ok=True)
        seen["candidates"].extend(new_entries)
        tmp = p.with_name(p.name + ".tmp")
        tmp.write_text(json.dumps(seen, indent=2))
        tmp.replace(p)


if __name__ == "__main__":
    main()
